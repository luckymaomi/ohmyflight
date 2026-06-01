# 锁班皇帝 - 并发实验版

from __future__ import annotations

import asyncio
import os
import platform
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from time import perf_counter

from colorama import Fore, Style, init
from openpyxl import Workbook, load_workbook
from playwright.async_api import TimeoutError as PlaywrightTimeoutError
from playwright.async_api import async_playwright


init()

LOGIN_URL = "https://ieb.csair.com/login"
INDEX_URL = "https://ieb.csair.com/index/index"
DEFAULT_CONCURRENCY = 20
MAX_ROUNDS = 20
TEXT_RUNNING = "运行管理"
TEXT_NON_PRODUCTION = "非生产任务"
TEXT_NON_PRODUCTION_ENTRY = "非生产任务录入"
TEXT_NEXT = "下一步"
TEXT_CONTINUE = "继续录入"
TEXT_NO_RELATED_INFO = "没有相关信息"
EXCEL_DATE_FORMAT = 'yyyy"年"m"月"d"日"'


LEAVE_TYPE_MAP = {
    "ALV_FD-飞行员公休（订座）": "ALV_FD",
    "RECU_LVE-健康疗养": "RECU_LVE",
    "RECU_LVE_R-康复疗养": "RECU_LVE_R",
    "MAT_FA_LVE-陪产假": "MAT_FA_LVE",
    "MAT_MO_LVE-产假": "MAT_MO_LVE",
    "PREGNANT-孕假": "PREGNANT",
    "PARENT_LVE-探亲假-探父母": "PARENT_LVE",
    "SPOUSE_LVE-探亲假-探配偶": "SPOUSE_LVE",
    "MARR_LVE-婚假": "MARR_LVE",
    "COMP_LVE-丧假": "COMP_LVE",
    "CHILD_LVE-育儿假": "CHILD_LVE",
    "INJURY_LVE-工伤假": "INJURY_LVE",
    "LWOP_LVE-其他（事假）": "LWOP_LVE",
    "UNPAID_LVE-无薪": "UNPAID_LVE",
    "HOUSE_LVE-搬家": "HOUSE_LVE",
    "BREED_LVE-哺乳假": "BREED_LVE",
    "PATERNITY-独生子女护理假": "PATERNITY",
    "BIRC_LVE-计划生育假": "BIRC_LVE",
    "REWARD_LVE-奖励": "REWARD_LVE",
    "PENALTY-停飞": "PENALTY",
    "PRD_LVE-经期假": "PRD_LVE",
    "GRD-地面班": "GRD",
    "GDO-地面休息": "GDO",
    "TRNG1-训练": "TRNG1",
    "BS_STUDY-业务学习": "BS_STUDY",
    "BUSINESS-公务": "BUSINESS",
    "GRD_ONDUTY-地面值班": "GRD_ONDUTY",
    "LG_STUDY-语言学习/考试": "LG_STUDY",
    "MEDL_CHK-体检_临床": "MEDL_CHK",
    "MEDL_PHLE-体检_抽血": "MEDL_PHLE",
    "MEDL_EET-体检_平板": "MEDL_EET",
    "MEDL_PSYC-体检_心理测试": "MEDL_PSYC",
    "MTG-会议": "MTG",
    "MTG_SF-安全讲评会": "MTG_SF",
    "DGET-危险品培训": "DGET",
    "EP-飞行人员应急复训": "EP",
    "CRM-CRM培训": "CRM",
    "T_SIM_INS-模拟机检查": "T_SIM_INS",
    "T_SIM_REC-模拟机复训": "T_SIM_REC",
    "T_SIM_INT-模拟机初始": "T_SIM_INT",
    "T_SIM_UPG-模拟机升级": "T_SIM_UPG",
    "T_SIM_CON-模拟机_转机型": "T_SIM_CON",
    "MAKEUP-补考": "MAKEUP",
    "BS_CONCL-飞行后讲评": "BS_CONCL",
    "BS_CHK-业务检查": "BS_CHK",
    "ADMN-管理任务": "ADMN",
    "SOCIAL-社会活动": "SOCIAL",
    "HANDBOOK-手册": "HANDBOOK",
    "POL_STUDY-政治学习": "POL_STUDY",
    "T/A-部门活动": "T/A",
}
LEAVE_CODE_TO_NAME = {value: key for key, value in LEAVE_TYPE_MAP.items()}
LEAVE_CODE_PATTERN = re.compile(
    r"(?<![A-Z0-9_/])("
    + "|".join(re.escape(code) for code in sorted(LEAVE_CODE_TO_NAME, key=len, reverse=True))
    + r")(?![A-Z0-9_/])"
)

RESULT_HEADERS = [
    "序号",
    "员工号",
    "姓名",
    "锁班类型",
    "开始日期",
    "结束日期",
    "处理状态",
    "锁班结果",
    "结果姓名",
    "结果锁班类型",
    "结果开始日期",
    "结果结束日期",
    "冲突",
    "备注",
    "员工号匹配",
    "姓名匹配",
    "日期匹配",
    "类型匹配",
    "处理时间",
]


@dataclass
class LockRecord:
    sequence: int
    employee_id: str
    name: str
    leave_type: str
    start_date: str
    end_date: str
    row_num: int = 0


@dataclass
class LockResult:
    sequence: int
    worker_id: int
    record: LockRecord
    status: str
    row: dict
    conflict: str
    remark: str
    entry_seconds: float | str
    fill_seconds: float | str
    submit_seconds: float | str
    total_seconds: float | str
    error: str = ""


@dataclass
class PortalResult:
    worker_id: int
    attempted_record: LockRecord
    status: str
    row: dict
    conflict: str
    remark: str
    fill_seconds: float | str
    submit_seconds: float | str
    total_seconds: float | str
    error: str = ""
    retryable: bool = True


def c_info(text: str) -> str:
    return f"{Fore.CYAN}{text}{Style.RESET_ALL}"


def c_ok(text: str) -> str:
    return f"{Fore.GREEN}{text}{Style.RESET_ALL}"


def c_err(text: str) -> str:
    return f"{Fore.RED}{text}{Style.RESET_ALL}"


def c_warn(text: str) -> str:
    return f"{Fore.YELLOW}{text}{Style.RESET_ALL}"


def c_hint(text: str) -> str:
    return f"{Fore.MAGENTA}{text}{Style.RESET_ALL}"


def beep_error() -> None:
    try:
        if platform.system() == "Windows":
            import winsound

            winsound.Beep(800, 300)
    except Exception:
        pass


def normalize_text(value) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_date(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    parts = re.split(r"[-/]", text)
    if len(parts) == 3:
        return f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
    return text


def normalize_employee_id(value) -> str:
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = normalize_text(value)
    match = re.search(r"\d{6}", text)
    return match.group(0) if match else text


def parse_whitelist(text: str) -> set[str]:
    all_nums = re.findall(r"\d{6}", re.sub(r"\D", " ", text))
    if all_nums:
        return set(all_nums)
    digits = re.sub(r"\D", "", text)
    return {digits[index : index + 6] for index in range(0, len(digits), 6) if len(digits[index : index + 6]) == 6}


def parse_leave_type(text: str) -> str:
    if not text:
        return ""
    text = str(text).strip()
    if text in LEAVE_CODE_TO_NAME:
        return text
    if text in LEAVE_TYPE_MAP:
        return LEAVE_TYPE_MAP[text]

    code_match = LEAVE_CODE_PATTERN.search(text)
    if code_match:
        return code_match.group(1)

    for key, code in sorted(LEAVE_TYPE_MAP.items(), key=lambda item: len(item[0]), reverse=True):
        if key in text:
            return code

    for key, code in sorted(LEAVE_TYPE_MAP.items(), key=lambda item: len(item[0]), reverse=True):
        name_part = key.split("-", 1)[1] if "-" in key else key
        if name_part in text or text in name_part:
            return code
    return ""


def leave_type_name(leave_type: str) -> str:
    display = LEAVE_CODE_TO_NAME.get(leave_type, leave_type or "")
    return display.split("-", 1)[1] if "-" in display else display


def same_day(left: str, right: str) -> bool:
    return normalize_date(left[:10]) == normalize_date(right[:10])


def is_row_number(value: str) -> bool:
    text = normalize_text(value)
    return bool(re.fullmatch(r"\d{1,4}", text))


def align_table_values(headers: list[str], values: list[str]) -> list[str]:
    if headers and values and len(values) == len(headers) + 1 and is_row_number(values[0]) and headers[0] != "序号":
        return values[1:]
    return values


def result_identity_problem(record: "LockRecord", row: dict) -> str:
    result_employee_id = row.get("员工号", "")
    if result_employee_id and result_employee_id != record.employee_id:
        return f"结果员工号不匹配: 输入{record.employee_id}, 页面{result_employee_id}"

    result_name = row.get("姓名", "")
    if result_name and record.name and result_name != record.name:
        return f"结果姓名不匹配: 输入{record.name}, 页面{result_name}"

    result_start = row.get("开始日期", "")
    result_end = row.get("结束日期", "")
    if result_start and not same_day(result_start, record.start_date):
        return f"结果开始日期不匹配: 输入{record.start_date}, 页面{result_start}"
    if result_end and not same_day(result_end, record.end_date):
        return f"结果结束日期不匹配: 输入{record.end_date}, 页面{result_end}"

    result_type = row.get("锁班类型", "")
    expected_type = leave_type_name(record.leave_type)
    if result_type and result_type != expected_type:
        return f"结果锁班类型不匹配: 输入{expected_type}, 页面{result_type}"
    return ""


def result_row_matches_record(row: dict, record: "LockRecord") -> bool:
    if no_related_info(row):
        return False

    row_text = row.get("_text", "")
    result_employee_id = row.get("员工号", "")
    result_name = row.get("姓名", "")
    result_start = row.get("开始日期", "")
    result_end = row.get("结束日期", "")
    result_type = row.get("锁班类型", "")
    expected_type = leave_type_name(record.leave_type)

    if result_employee_id:
        if result_employee_id != record.employee_id:
            return False
    elif record.employee_id not in row_text:
        return False

    if record.name:
        if result_name:
            if result_name != record.name:
                return False
        elif record.name not in row_text:
            return False

    if result_type and result_type != expected_type:
        return False
    if not result_type and expected_type not in row_text:
        return False

    if result_start:
        if not same_day(result_start, record.start_date):
            return False
    elif record.start_date not in row_text:
        return False

    if result_end:
        if not same_day(result_end, record.end_date):
            return False
    elif record.end_date not in row_text:
        return False

    return True


def parse_single_record(line: str, sequence: int) -> LockRecord | None:
    employee_match = re.search(r"\b(\d{6})\b", line)
    if not employee_match:
        return None
    dates = re.findall(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}", line)
    if not dates:
        return None
    employee_id = employee_match.group(1)
    name_match = re.search(r"\d{6}\s*([\u4e00-\u9fa5]{2,4})", line)
    leave_type = parse_leave_type(line)
    return LockRecord(
        sequence=sequence,
        employee_id=employee_id,
        name=name_match.group(1) if name_match else "",
        leave_type=leave_type,
        start_date=normalize_date(dates[0]),
        end_date=normalize_date(dates[1] if len(dates) > 1 else dates[0]),
    )


def read_multiline_records(whitelist: set[str] | None = None) -> list[LockRecord]:
    print(c_hint("请粘贴锁班记录，每行一条，输入 ok 开始，输入 c 取消："))
    lines = []
    while True:
        line = input()
        lowered = line.strip().lower()
        if lowered == "c":
            return []
        if lowered == "ok":
            break
        if line.strip():
            lines.append(line.strip())

    records = []
    for index, line in enumerate(lines, start=1):
        record = parse_single_record(line, len(records) + 1)
        if whitelist and record and record.employee_id not in whitelist:
            continue
        if record and record.leave_type:
            records.append(record)
        else:
            print(c_warn(f"第{index}行未识别，已跳过: {line[:60]}"))
    return records


def read_excel_records(filepath: str, whitelist: set[str] | None = None) -> tuple[list[LockRecord], list[str]]:
    records = []
    errors = []
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(row):
            continue
        employee_id = normalize_employee_id(row[0] if len(row) > 0 else "")
        name = normalize_text(row[1] if len(row) > 1 else "")
        leave_type = parse_leave_type(normalize_text(row[2] if len(row) > 2 else ""))
        start_date = normalize_date(row[3] if len(row) > 3 else "")
        end_date = normalize_date(row[4] if len(row) > 4 and row[4] else start_date)

        if not re.fullmatch(r"\d{6}", employee_id):
            errors.append(f"第{row_num}行员工号无效: {employee_id}")
            continue
        if not leave_type:
            errors.append(f"第{row_num}行锁班类型无效")
            continue
        if not start_date:
            errors.append(f"第{row_num}行开始日期无效")
            continue
        if whitelist and employee_id not in whitelist:
            continue

        records.append(
            LockRecord(
                sequence=len(records) + 1,
                employee_id=employee_id,
                name=name,
                leave_type=leave_type,
                start_date=start_date,
                end_date=end_date or start_date,
                row_num=row_num,
            )
        )
    wb.close()
    return records, errors


def create_result_excel(label: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_label = re.sub(r'[\\/:*?"<>|]+', "_", label or "super_lock")
    output_file = os.path.abspath(f"{safe_label}_super结果_{timestamp}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "results"
    ws.append(RESULT_HEADERS)
    wb.save(output_file)
    wb.close()
    return output_file


def append_result_excel(output_file: str, result: LockResult) -> None:
    record = result.record
    row = result.row or {}
    result_employee_id = row.get("员工号", "")
    result_name = row.get("姓名", "")
    result_start = row.get("开始日期", "")
    result_end = row.get("结束日期", "")
    result_type = row.get("锁班类型", "")
    expected_type_name = leave_type_name(record.leave_type)

    employee_match = (not result_employee_id) or result_employee_id == record.employee_id
    name_match = (not result_name) or (not record.name) or result_name == record.name
    date_match = (not result_start or same_day(result_start, record.start_date)) and (
        not result_end or same_day(result_end, record.end_date)
    )
    type_match = (not result_type) or result_type == expected_type_name

    wb = load_workbook(output_file)
    ws = wb.active
    ws.append(
        [
            result.sequence,
            record.employee_id,
            record.name,
            expected_type_name,
            record.start_date,
            record.end_date,
            result.status,
            row.get("锁班结果") or row.get("锁班状态") or result.status,
            row.get("姓名", ""),
            result_type,
            result_start,
            result_end,
            result.conflict,
            "" if result.status == "成功" else (result.remark or result.error),
            "是" if employee_match else "否",
            "是" if name_match else "否",
            "是" if date_match else "否",
            "是" if type_match else "否",
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ]
    )
    wb.save(output_file)
    wb.close()


async def writer(output_file: str, result_queue: asyncio.Queue, total: int) -> None:
    written = 0
    while written < total:
        result = await result_queue.get()
        append_result_excel(output_file, result)
        written += 1
        status_text = f"[{written}/{total}] worker {result.worker_id} {result.record.employee_id} {result.status}"
        if result.remark:
            status_text += f" {result.remark}"
        print(c_ok(status_text) if result.status == "成功" else c_warn(status_text))
        result_queue.task_done()


async def click_link(page, name: str) -> None:
    link = page.get_by_role("link", name=name)
    await link.wait_for(timeout=20000)
    await link.click()
    await page.wait_for_timeout(500)


async def open_entry_page(page) -> None:
    await page.goto(INDEX_URL)
    await page.wait_for_load_state("networkidle")
    await page.get_by_text(TEXT_RUNNING).nth(1).wait_for(timeout=20000)
    await page.get_by_text(TEXT_RUNNING).nth(1).click()
    await page.wait_for_timeout(700)
    await click_link(page, TEXT_NON_PRODUCTION)
    await click_link(page, TEXT_NON_PRODUCTION_ENTRY)
    await page.locator("#mainContent").wait_for(timeout=20000)
    await page.locator("#mainContent").click()
    await page.wait_for_timeout(800)


async def clear_form(page) -> None:
    await page.locator("#showIdshowNonproductionTaskImportPage").fill("")
    await page.locator("#lockStartTime").fill("")
    await page.locator("#lockEndTime").fill("")
    reason_input = page.locator("#lockReasonTxt")
    if await reason_input.count() > 0:
        try:
            await reason_input.fill("")
        except Exception:
            pass


async def fill_form(page, record: LockRecord, reason_text: str) -> None:
    await clear_form(page)
    employee_input = page.locator("#showIdshowNonproductionTaskImportPage")
    await employee_input.click()
    await employee_input.fill("")
    await employee_input.type(record.employee_id, delay=0)
    await page.evaluate(
        """
        const input = document.querySelector('#showIdshowNonproductionTaskImportPage');
        if (input) {
            input.dispatchEvent(new Event('blur', { bubbles: true }));
            input.dispatchEvent(new Event('change', { bubbles: true }));
        }
        """
    )
    await page.wait_for_timeout(250)
    await page.evaluate(
        """(leaveType) => {
            const select = document.querySelector('#lockType');
            if (select) {
                select.focus();
                select.value = leaveType;
                select.dispatchEvent(new MouseEvent('mousedown', { bubbles: true }));
                select.dispatchEvent(new MouseEvent('click', { bubbles: true }));
                select.dispatchEvent(new Event('input', { bubbles: true }));
                select.dispatchEvent(new Event('change', { bubbles: true }));
                select.dispatchEvent(new Event('blur', { bubbles: true }));
            }
        }""",
        record.leave_type,
    )
    await page.wait_for_timeout(100)
    await page.locator("#lockStartTime").fill(record.start_date)
    await page.locator("#lockEndTime").fill(record.end_date)

    reason_input = page.locator("#lockReasonTxt")
    if reason_text and await reason_input.count() > 0:
        await reason_input.click()
        await reason_input.fill("")
        await reason_input.fill(reason_text)


async def table_headers(page, selector: str) -> list[str]:
    cells = page.locator(selector)
    headers = []
    for index in range(await cells.count()):
        try:
            text = normalize_text(await cells.nth(index).inner_text(timeout=1000))
        except Exception:
            text = ""
        if text:
            headers.append(text)
    return headers


async def table_rows(page, selector: str, headers: list[str]) -> list[dict]:
    row_locs = page.locator(selector)
    rows = []
    for row_index in range(await row_locs.count()):
        row = row_locs.nth(row_index)
        cell_locs = row.locator("td")
        values = []
        for cell_index in range(await cell_locs.count()):
            try:
                values.append(normalize_text(await cell_locs.nth(cell_index).inner_text(timeout=1000)))
            except Exception:
                values.append("")

        if not values:
            try:
                values = [normalize_text(await row.inner_text(timeout=1000))]
            except Exception:
                values = []
        else:
            values = align_table_values(headers, values)

        row_map = {}
        for index, value in enumerate(values):
            key = headers[index] if index < len(headers) else f"列{index + 1}"
            row_map[key] = value
        row_map["_text"] = " | ".join(value for value in values if value)
        rows.append(row_map)
    return rows


def no_related_info(row: dict) -> bool:
    return TEXT_NO_RELATED_INFO in row.get("_text", "")


def first_matching_row(rows: list[dict], record: LockRecord) -> dict | None:
    for row in rows:
        if result_row_matches_record(row, record):
            return row
    return None


async def read_submit_result(page, record: LockRecord) -> tuple[str, dict, str]:
    result_headers = await table_headers(page, "#showNonproductionTaskImportResultPage1 th")
    conflict_headers = await table_headers(page, "#showNonproductionTaskImportResultPage2 th")
    result_rows = await table_rows(page, "#showNonproductionTaskImportResultPage1 tbody.list tr", result_headers)
    conflict_rows = await table_rows(page, "#showNonproductionTaskImportResultPage2 tbody.list tr", conflict_headers)

    real_results = [row for row in result_rows if not no_related_info(row)]
    real_conflicts = [row for row in conflict_rows if not no_related_info(row)]
    matching_result = first_matching_row(real_results, record)
    matching_conflict = first_matching_row(real_conflicts, record)

    if real_conflicts:
        if not matching_conflict:
            first_text = real_conflicts[0].get("_text", "") if real_conflicts else ""
            return "异常", {}, f"冲突结果未匹配当前人员: {record.employee_id} {record.name}; 首行{first_text[:120]}"
        conflict = matching_conflict.get("冲突") or matching_conflict.get("_text", "")
        return "冲突", matching_conflict, conflict
    if real_results and any(no_related_info(row) for row in conflict_rows):
        if not matching_result:
            first_text = real_results[0].get("_text", "") if real_results else ""
            return "异常", {}, f"提交结果未匹配当前人员: {record.employee_id} {record.name}; 首行{first_text[:120]}"
        return "成功", matching_result, ""
    if not real_results:
        return "失败", {}, "查询结果为空"
    if not matching_result:
        first_text = real_results[0].get("_text", "") if real_results else ""
        return "异常", {}, f"提交结果未匹配当前人员: {record.employee_id} {record.name}; 首行{first_text[:120]}"
    note = "; ".join(row.get("_text", "") for row in conflict_rows if row.get("_text")) or "未知结果"
    return "失败", matching_result, note


async def read_any_submit_result(page) -> tuple[str, dict, str]:
    result_headers = await table_headers(page, "#showNonproductionTaskImportResultPage1 th")
    conflict_headers = await table_headers(page, "#showNonproductionTaskImportResultPage2 th")
    result_rows = await table_rows(page, "#showNonproductionTaskImportResultPage1 tbody.list tr", result_headers)
    conflict_rows = await table_rows(page, "#showNonproductionTaskImportResultPage2 tbody.list tr", conflict_headers)

    real_results = [row for row in result_rows if not no_related_info(row)]
    real_conflicts = [row for row in conflict_rows if not no_related_info(row)]

    if real_conflicts:
        row = real_conflicts[0]
        conflict = row.get("冲突") or row.get("_text", "")
        return "冲突", row, conflict
    if real_results and any(no_related_info(row) for row in conflict_rows):
        return "成功", real_results[0], ""
    if not real_results:
        return "失败", {}, "查询结果为空"
    note = "; ".join(row.get("_text", "") for row in conflict_rows if row.get("_text")) or "未知结果"
    return "失败", real_results[0], note


async def wait_for_submit_result(page, timeout_ms: int = 90000) -> tuple[str, dict, str]:
    deadline = perf_counter() + timeout_ms / 1000
    last_status = "失败"
    last_row: dict = {}
    last_remark = "等待结果超时"

    while perf_counter() < deadline:
        status, row, remark = await read_any_submit_result(page)
        last_status, last_row, last_remark = status, row, remark
        if row or status == "失败":
            return status, row, remark
        await page.wait_for_timeout(500)

    return last_status, last_row, last_remark


async def has_result_page(page) -> bool:
    try:
        return await page.get_by_role("button", name=TEXT_CONTINUE).count() > 0
    except Exception:
        return False


async def read_info_dialog(page) -> str:
    dialog = page.locator(".ui-dialog:visible").filter(has=page.locator("#info-dialog")).first
    try:
        await dialog.wait_for(timeout=1200)
    except Exception:
        return ""

    try:
        content = normalize_text(await dialog.locator("#info-dialog").inner_text(timeout=1000))
        full_text = normalize_text(await dialog.inner_text(timeout=1000))
        text = content or full_text
    except Exception:
        text = "提示信息弹窗"

    return text or "提示信息弹窗"


async def wait_for_result_or_dialog(page, timeout_ms: int = 90000) -> tuple[str, dict, str, bool]:
    deadline = perf_counter() + timeout_ms / 1000
    while perf_counter() < deadline:
        dialog_text = await read_info_dialog(page)
        if dialog_text:
            return "异常", {}, dialog_text, False
        if await has_result_page(page):
            status, row, remark = await wait_for_submit_result(page)
            return status, row, remark, True
        await page.wait_for_timeout(300)
    return "失败", {}, "等待结果页或提示信息弹窗超时", True


async def go_back_to_form(page) -> None:
    try:
        await page.get_by_role("button", name=TEXT_CONTINUE).click(timeout=3000)
        await page.locator("#showIdshowNonproductionTaskImportPage").wait_for(timeout=10000)
    except Exception:
        await open_entry_page(page)


def row_summary(row: dict) -> str:
    return row.get("_text", "")[:180] if row else ""


def portal_result_to_lock_result(portal_result: PortalResult, record: LockRecord, status: str, remark: str) -> LockResult:
    return LockResult(
        sequence=record.sequence,
        worker_id=portal_result.worker_id,
        record=record,
        status=status,
        row=portal_result.row,
        conflict=remark if status == "冲突" else "",
        remark=remark,
        entry_seconds="",
        fill_seconds=portal_result.fill_seconds,
        submit_seconds=portal_result.submit_seconds,
        total_seconds=portal_result.total_seconds,
        error=portal_result.error,
    )


def match_portal_results(
    pending_records: list[LockRecord],
    portal_results: list[PortalResult],
) -> tuple[list[LockResult], list[LockRecord], dict[int, str]]:
    unmatched_records = pending_records[:]
    matched_results: list[LockResult] = []
    attempt_notes: dict[int, str] = {}

    for portal_result in portal_results:
        attempted = portal_result.attempted_record
        note = portal_result.remark or portal_result.error or row_summary(portal_result.row) or portal_result.status
        attempt_notes[attempted.sequence] = note

        if portal_result.error or not portal_result.row:
            if not portal_result.retryable:
                if attempted in unmatched_records:
                    unmatched_records.remove(attempted)
                matched_results.append(
                    LockResult(
                        sequence=attempted.sequence,
                        worker_id=portal_result.worker_id,
                        record=attempted,
                        status="异常",
                        row={},
                        conflict="",
                        remark=portal_result.remark or portal_result.error,
                        entry_seconds="",
                        fill_seconds=portal_result.fill_seconds,
                        submit_seconds=portal_result.submit_seconds,
                        total_seconds=portal_result.total_seconds,
                        error=portal_result.error,
                    )
                )
            continue

        matched_index = None
        for index, record in enumerate(unmatched_records):
            if result_row_matches_record(portal_result.row, record):
                matched_index = index
                break

        if matched_index is None:
            continue

        matched_record = unmatched_records.pop(matched_index)
        identity_error = result_identity_problem(matched_record, portal_result.row)
        if identity_error:
            attempt_notes[matched_record.sequence] = identity_error
            unmatched_records.append(matched_record)
            continue

        matched_results.append(
            portal_result_to_lock_result(
                portal_result,
                matched_record,
                portal_result.status,
                portal_result.conflict or portal_result.remark,
            )
        )

    return matched_results, unmatched_records, attempt_notes


async def process_record(
    page,
    record: LockRecord,
    worker_id: int,
    reason_text: str,
) -> PortalResult:
    started = perf_counter()
    try:
        await fill_form(page, record, reason_text)
        fill_done = perf_counter()
        await page.get_by_role("button", name=TEXT_NEXT).wait_for(timeout=20000)
        await page.get_by_role("button", name=TEXT_NEXT).click()
        status, row, remark, retryable = await wait_for_result_or_dialog(page)
        submit_done = perf_counter()
        if retryable:
            await go_back_to_form(page)

        return PortalResult(
            worker_id=worker_id,
            attempted_record=record,
            status=status,
            row=row,
            conflict=remark if status == "冲突" else "",
            remark=remark,
            fill_seconds=round(fill_done - started, 3),
            submit_seconds=round(submit_done - fill_done, 3),
            total_seconds=round(perf_counter() - started, 3),
            error=remark if not retryable else "",
            retryable=retryable,
        )
    except Exception as exc:
        try:
            await go_back_to_form(page)
        except Exception:
            pass
        return PortalResult(
            worker_id=worker_id,
            attempted_record=record,
            status="异常",
            row={},
            conflict="",
            remark="",
            fill_seconds="",
            submit_seconds="",
            total_seconds=round(perf_counter() - started, 3),
            error=str(exc).splitlines()[0],
        )


async def worker_loop(
    context,
    worker_id: int,
    task_queue: asyncio.Queue,
    portal_result_queue: asyncio.Queue,
    reason_text: str,
) -> None:
    page = await context.new_page()
    page.set_default_timeout(30000)
    entry_started = perf_counter()
    try:
        await open_entry_page(page)
        entry_seconds = round(perf_counter() - entry_started, 3)
        while True:
            record = await task_queue.get()
            if record is None:
                task_queue.task_done()
                break

            result = await process_record(page, record, worker_id, reason_text)
            await portal_result_queue.put(result)
            if not result.retryable:
                try:
                    await page.close()
                except Exception:
                    pass
                page = await context.new_page()
                page.set_default_timeout(30000)
                await open_entry_page(page)
            task_queue.task_done()
    except Exception as exc:
        while True:
            try:
                record = task_queue.get_nowait()
            except asyncio.QueueEmpty:
                break
            if record is not None:
                await portal_result_queue.put(
                    PortalResult(
                        worker_id=worker_id,
                        attempted_record=record,
                        status="异常",
                        row={},
                        conflict="",
                        remark="",
                        fill_seconds="",
                        submit_seconds="",
                        total_seconds="",
                        error=f"worker初始化失败: {str(exc).splitlines()[0]}",
                    )
                )
            task_queue.task_done()
    finally:
        try:
            await page.close()
        except Exception:
            pass


async def run_round(
    context,
    records: list[LockRecord],
    concurrency: int,
    reason_text: str,
) -> list[PortalResult]:
    task_queue: asyncio.Queue = asyncio.Queue()
    portal_result_queue: asyncio.Queue = asyncio.Queue()
    worker_count = min(concurrency, len(records))

    for record in records:
        await task_queue.put(record)
    for _ in range(worker_count):
        await task_queue.put(None)

    worker_tasks = [
        asyncio.create_task(worker_loop(context, worker_id, task_queue, portal_result_queue, reason_text))
        for worker_id in range(1, worker_count + 1)
    ]

    await task_queue.join()
    await asyncio.gather(*worker_tasks)

    portal_results = []
    while not portal_result_queue.empty():
        portal_results.append(await portal_result_queue.get())
        portal_result_queue.task_done()
    return portal_results


def unmatched_lock_result(record: LockRecord, note: str) -> LockResult:
    return LockResult(
        sequence=record.sequence,
        worker_id=0,
        record=record,
        status="未匹配",
        row={},
        conflict="",
        remark=f"重跑{MAX_ROUNDS}轮后仍未匹配: {note or '无页面结果'}",
        entry_seconds="",
        fill_seconds="",
        submit_seconds="",
        total_seconds="",
    )


def whitelist_status(whitelist: set[str] | None) -> str:
    if whitelist:
        return c_ok(f"白名单:{len(whitelist)}人")
    return c_warn("白名单:无")


def collect_records(whitelist: set[str] | None = None) -> list[LockRecord]:
    print(whitelist_status(whitelist))
    mode = input(c_hint("选择输入方式：1粘贴 2读取Excel（默认2）: ")).strip() or "2"
    if mode == "1":
        return read_multiline_records(whitelist)

    filepath = input(c_hint("请输入Excel文件路径: ")).strip().strip('"').strip("'")
    if not filepath or not os.path.exists(filepath):
        print(c_err("Excel文件不存在"))
        return []
    records, errors = read_excel_records(filepath, whitelist)
    for error in errors:
        print(c_warn(error))
    return records


def format_reason_preview(reason_text: str, limit: int = 18) -> str:
    preview = reason_text.replace("\n", " / ")
    if len(preview) > limit:
        return preview[:limit] + "..."
    return preview


def read_multiline(prompt: str, confirm_key: str = "ok", cancel_key: str = "c") -> str | None:
    print(prompt)
    lines: list[str] = []
    while True:
        line = input()
        if line.lower() == cancel_key:
            return None
        if line.lower() == confirm_key:
            break
        if line:
            lines.append(line)
    if not lines:
        return None
    return "\n".join(lines)


def set_whitelist() -> set[str] | None:
    text = read_multiline(c_hint("请粘贴员工号列表(输入ok确认,c取消):"), "ok", "c")
    if text is None:
        print(c_warn("已取消"))
        return None
    whitelist = parse_whitelist(text)
    if not whitelist:
        print(c_err("未识别到有效员工号"))
        return None
    print(c_ok(f"已设置白名单,共{len(whitelist)}人"))
    return whitelist


def set_common_reason() -> str | None:
    text = read_multiline(c_hint("请粘贴统一备注(输入OK确认,c取消):"), "ok", "c")
    if text is None:
        print(c_warn("本次不填写备注"))
        return None
    print(c_ok(f"已设置统一备注: {format_reason_preview(text, limit=30)}"))
    return text


async def prepare_entry_page(context):
    page = await context.new_page()
    page.set_default_timeout(30000)
    try:
        await page.goto(LOGIN_URL)
        await page.wait_for_load_state("networkidle")
        await page.locator("#scanLogin").wait_for()
        await page.locator("#scanLogin").click()
        print(c_info("请扫码登录..."))
        try:
            await page.wait_for_url("**/index/**", timeout=240000)
            print(c_ok("登录成功"))
        except PlaywrightTimeoutError:
            print(c_warn("未检测到自动跳转，将继续尝试进入首页。"))
    except Exception as exc:
        print(c_warn(f"自动登录流程未完成: {str(exc).splitlines()[0]}"))
        print(c_warn("请手动完成登录"))
        input(c_hint("登录完成后按回车继续..."))

    try:
        print(c_info("正在进入非生产任务录入页面..."))
        await open_entry_page(page)
        print(c_ok("已进入非生产任务录入页面"))
    except Exception as exc:
        print(c_warn(f"自动导航失败: {str(exc).splitlines()[0]}"))
        print(c_warn("请手动进入非生产任务录入页面"))
        input(c_hint("准备好后按回车继续..."))
    return page


def preview_records(records: list[LockRecord], concurrency: int) -> bool:
    print(c_ok(f"共 {len(records)} 条记录，并发 {concurrency}。"))
    for record in records[:20]:
        print(f"{record.sequence}. {record.employee_id} {record.name} {record.leave_type} {record.start_date}~{record.end_date}")
    if len(records) > 20:
        print(c_warn(f"仅预览前20条，剩余 {len(records) - 20} 条不逐条显示。"))
    confirm = input(c_hint("输入 y 开始并发锁班，其他键退出: ")).strip().lower()
    return confirm == "y"


async def run(concurrency: int, reason_text: str, browser_path: str | None, whitelist: set[str] | None) -> str:
    async with async_playwright() as playwright:
        browser = await playwright.chromium.launch(headless=False, executable_path=browser_path or None, slow_mo=0)
        context = await browser.new_context(ignore_https_errors=True)
        ready_page = None
        try:
            ready_page = await prepare_entry_page(context)
            print(c_ok("开始导入锁班数据"))
            records = collect_records(whitelist)
            if not records:
                print(c_err("没有可处理记录"))
                return ""
            if not preview_records(records, concurrency):
                return ""
            output_file = create_result_excel("并发锁班")
            await ready_page.close()
            ready_page = None

            result_queue: asyncio.Queue = asyncio.Queue()
            writer_task = asyncio.create_task(writer(output_file, result_queue, len(records)))

            pending_records = records[:]
            last_notes: dict[int, str] = {}
            written_sequences: set[int] = set()

            for round_index in range(1, MAX_ROUNDS + 1):
                if not pending_records:
                    break
                print(c_info(f"第{round_index}轮：待处理 {len(pending_records)} 条，并发 {min(concurrency, len(pending_records))}。"))
                portal_results = await run_round(context, pending_records, concurrency, reason_text)
                matched_results, pending_records, attempt_notes = match_portal_results(pending_records, portal_results)
                last_notes.update(attempt_notes)
                print(c_info(f"第{round_index}轮：匹配 {len(matched_results)} 条，剩余 {len(pending_records)} 条。"))

                for result in matched_results:
                    if result.sequence not in written_sequences:
                        await result_queue.put(result)
                        written_sequences.add(result.sequence)

            for record in pending_records:
                if record.sequence not in written_sequences:
                    await result_queue.put(unmatched_lock_result(record, last_notes.get(record.sequence, "")))
                    written_sequences.add(record.sequence)

            await result_queue.join()
            await writer_task
        finally:
            if ready_page is not None:
                try:
                    await ready_page.close()
                except Exception:
                    pass
            await context.close()
            await browser.close()
    return output_file


def main() -> int:
    print(c_info("锁班皇帝 - 并发实验版"))
    print(c_warn("实验版不会替换 app.py。并发越高越需要核对结果Excel里的匹配校验。"))
    browser_path = input(c_hint("浏览器路径(回车用默认): ")).strip() or None
    if browser_path:
        print(c_ok(f"使用指定浏览器: {browser_path}"))
    else:
        print(c_ok("使用默认浏览器"))
    whitelist = None
    use_wl = input(c_hint("是否预设白名单?(y/n): ")).strip().lower()
    if use_wl == "y":
        whitelist = set_whitelist()
    else:
        print(c_ok("不设置白名单,处理所有员工"))
    reason_text = ""
    use_reason = input(c_hint("是否填写统一备注?(y/n): ")).strip().lower()
    if use_reason == "y":
        reason_text = set_common_reason() or ""
    else:
        print(c_ok("本次不填写备注"))
    raw_concurrency = input(c_hint(f"并发数(默认 {DEFAULT_CONCURRENCY}): ")).strip()
    concurrency = int(raw_concurrency) if raw_concurrency.isdigit() and int(raw_concurrency) > 0 else DEFAULT_CONCURRENCY

    try:
        output_file = asyncio.run(run(concurrency, reason_text, browser_path, whitelist))
        if output_file:
            print(c_ok(f"处理完成，结果Excel: {output_file}"))
        return 0
    except KeyboardInterrupt:
        print(c_warn("收到中断信号，已停止。"))
        return 130


if __name__ == "__main__":
    sys.exit(main())
