# 飞行经历/左座经历起落数查询助手 - 批量模式（统信浏览器版本）

import re
import os
from datetime import datetime, date
from typing import Optional
from colorama import init, Fore, Style
from playwright.sync_api import sync_playwright
from openpyxl import Workbook, load_workbook

init()

EXCEL_DATE_FORMAT = 'yyyy"年"m"月"d"日"'
QUERY_TYPE_OPTIONS = {
    "1": ("飞行经历", 8),
    "2": ("左座经历", 11),
}
INPUT_MODE_OPTIONS = {
    "1": "读取Excel文件",
    "2": "直接粘贴数据",
}


def c_info(text):
    return f"{Fore.CYAN}{text}{Style.RESET_ALL}"

def c_ok(text):
    return f"{Fore.GREEN}{text}{Style.RESET_ALL}"

def c_err(text):
    return f"{Fore.RED}{text}{Style.RESET_ALL}"

def c_warn(text):
    return f"{Fore.YELLOW}{text}{Style.RESET_ALL}"

def c_hint(text):
    return f"{Fore.MAGENTA}{text}{Style.RESET_ALL}"


def normalize_date(date_str: str) -> str:
    """把各种日期格式统一转成YYYY/MM/DD"""
    parts = re.split(r'[-/]', str(date_str))
    if len(parts) == 3:
        year, month, day = parts
        return f"{year}/{month.zfill(2)}/{day.zfill(2)}"
    return date_str


def parse_excel_date(value):
    """把文本或日期对象转成可写入 Excel 的 datetime 值"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.replace(hour=0, minute=0, second=0, microsecond=0)
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())

    text = normalize_date(str(value).strip())
    if not text:
        return None

    try:
        return datetime.strptime(text, "%Y/%m/%d")
    except ValueError:
        return None


def format_query_date(value) -> Optional[str]:
    """统一查询用日期字符串格式"""
    excel_date = parse_excel_date(value)
    if excel_date:
        return excel_date.strftime("%Y/%m/%d")
    if value is None:
        return None
    text = str(value).strip()
    return normalize_date(text) if text else None


def same_query_date(left, right) -> bool:
    """比较两个查询日期是否为同一天"""
    left_date = parse_excel_date(left)
    right_date = parse_excel_date(right)
    return bool(left_date and right_date and left_date.date() == right_date.date())


def write_excel_date_cell(ws, row_num: int, column_num: int, value) -> None:
    """把日期写成真正的 Excel 日期单元格"""
    excel_date = parse_excel_date(value)
    if not excel_date:
        return
    cell = ws.cell(row=row_num, column=column_num)
    cell.value = excel_date
    cell.number_format = EXCEL_DATE_FORMAT


def parse_single_record(text: str) -> dict:
    """解析单条记录"""
    result = {
        "员工号": None,
        "姓名": None,
        "开始日期": None,
        "结束日期": None,
        "开始日期值": None,
        "结束日期值": None,
    }
    
    # 提取员工号
    emp = re.search(r'\b(\d{6})\b', text)
    if emp:
        result["员工号"] = emp.group(1)
    
    # 提取姓名（紧跟员工号后的中文）
    name = re.search(r'\d{6}\s*([\u4e00-\u9fa5]{2,4})', text)
    if name:
        result["姓名"] = name.group(1)
    
    # 提取日期
    dates = re.findall(r'\d{4}[-/]\d{1,2}[-/]\d{1,2}', text)
    if dates:
        start_date = format_query_date(dates[0])
        end_date = format_query_date(dates[1]) if len(dates) > 1 else start_date
        result["开始日期"] = start_date
        result["结束日期"] = end_date
        result["开始日期值"] = parse_excel_date(start_date)
        result["结束日期值"] = parse_excel_date(end_date)
    
    return result


def split_continuous_text(text: str) -> list:
    """把连续粘贴的文本按员工号切分成多条记录"""
    parts = re.split(r'(?=\d{6}[\u4e00-\u9fa5])', text)
    return [p.strip() for p in parts if p.strip() and re.search(r'\d{6}', p)]


def parse_batch_input(text: str) -> tuple:
    """解析批量输入"""
    records = []
    errors = []
    
    # 先按换行分，如果只有一行且很长，尝试按员工号切分
    lines = [line.strip() for line in text.strip().split('\n') if line.strip()]
    if len(lines) == 1 and len(lines[0]) > 100:
        lines = split_continuous_text(lines[0])
    
    for i, line in enumerate(lines, 1):
        record = parse_single_record(line)
        
        if not record["员工号"]:
            errors.append(f"第{i}条: 未识别员工号 [{line[:50]}]")
            continue
        
        if not record["开始日期"]:
            errors.append(f"第{i}条: 未识别日期 [{line[:50]}]")
            continue
        
        records.append(record)
    
    return records, errors


def read_multiline(prompt, confirm_key='ok', cancel_key='c'):
    """读取多行输入"""
    print(prompt)
    lines = []
    while True:
        line = input()
        if line.lower() == cancel_key:
            return None
        if line.lower() == confirm_key:
            break
        if line:
            lines.append(line)
    if not lines:
        return None
    return '\n'.join(lines)


def parse_excel_file(filepath: str) -> tuple:
    """解析Excel文件，返回(records, errors)"""
    records = []
    errors = []
    
    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue
            
            emp_id = str(int(row[0])).strip() if row[0] else None
            name = str(row[1]).strip() if len(row) > 1 and row[1] else None
            start_date = row[2] if len(row) > 2 else None
            end_date = row[3] if len(row) > 3 else None
            
            if not emp_id or not re.match(r'^\d{6}$', emp_id):
                if emp_id and emp_id != 'None':
                    errors.append(f"第{row_num}行: 员工号格式错误 [{emp_id}]")
                continue
            
            start_value = parse_excel_date(start_date)
            end_value = parse_excel_date(end_date) if end_date else start_value
            start = format_query_date(start_value)
            end = format_query_date(end_value) if end_value else start
            
            if not start:
                errors.append(f"第{row_num}行: 日期格式错误")
                continue
            
            records.append({
                "员工号": emp_id,
                "姓名": name,
                "开始日期": start,
                "结束日期": end or start,
                "开始日期值": start_value,
                "结束日期值": end_value or start_value,
                "行号": row_num
            })
        
        wb.close()
    except Exception as e:
        errors.append(f"读取Excel失败: {e}")
    
    return records, errors


def select_query_type(current_query=None):
    """选择查询类型，返回(经历名称, 列索引)；输入q返回None"""
    print(c_hint("\n请选择查询类型:"))
    print("1. 飞行经历")
    print("2. 左座经历")
    prompt = "请选择(1/2"
    if current_query:
        prompt += f"，回车沿用当前[{current_query[0]}]"
    prompt += "，q退出): "

    while True:
        choice = input(c_hint(prompt)).strip().lower()
        if not choice and current_query:
            return current_query
        if choice == "q":
            return None
        if choice in QUERY_TYPE_OPTIONS:
            return QUERY_TYPE_OPTIONS[choice]
        print(c_warn("无效选择，请重新输入"))


def select_input_mode(current_mode=None):
    """选择输入方式，返回1或2；输入q返回None"""
    print(c_hint("\n请选择输入方式:"))
    print("1. 读取Excel文件")
    print("2. 直接粘贴数据")
    prompt = "请选择(1/2"
    if current_mode:
        prompt += f"，回车沿用当前[{INPUT_MODE_OPTIONS[current_mode]}]"
    prompt += "，q退出): "

    while True:
        choice = input(c_hint(prompt)).strip().lower()
        if not choice and current_mode:
            return current_mode
        if choice == "q":
            return None
        if choice in INPUT_MODE_OPTIONS:
            return choice
        print(c_warn("无效选择，请重新输入"))


def write_to_excel(
    filepath: str,
    row_num: int,
    exp_value: str,
    landing_count: str,
    exp_label: str,
    start_date_value=None,
    end_date_value=None,
):
    """实时写入Excel"""
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        
        write_excel_date_cell(ws, row_num, 3, start_date_value)
        write_excel_date_cell(ws, row_num, 4, end_date_value)

        # 写入表头与数据：经历（E列）和起落总数（F列）
        ws.cell(row=1, column=5, value=exp_label)
        ws.cell(row=1, column=6, value="起落总数")
        ws.cell(row=row_num, column=5, value=exp_value)
        ws.cell(row=row_num, column=6, value=landing_count)
        
        wb.save(filepath)
        wb.close()
        return True
    except Exception as e:
        print(c_err(f"写入Excel失败: {e}"))
        return False


def fill_date(page, date_str):
    """填写日期到日期选择器 - 统信浏览器版本"""
    parts = date_str.split('/')
    if len(parts) != 3:
        raise ValueError(f"日期格式错误: {date_str}")
    
    year, month, day = parts
    
    # 月份映射 - 注意11月和12月的特殊格式
    month_map = {
        '01': '一月', '02': '二月', '03': '三月', '04': '四月',
        '05': '五月', '06': '六月', '07': '七月', '08': '八月',
        '09': '九月', '10': '十月', '11': '十一', '12': '十二'
    }
    
    month_cn = month_map.get(month, month)
    day_str = str(int(day))  # 去掉前导0
    
    # 等待iframe加载
    page.wait_for_timeout(300)
    
    # 使用frame_locator代替locator().content_frame
    frame = page.frame_locator("iframe >> nth=2")
    
    # 1. 点击年份输入框
    frame.get_by_role("textbox").nth(1).click()
    page.wait_for_timeout(100)
    
    # 2. 选择年份
    frame.get_by_role("cell", name=year, exact=True).click()
    page.wait_for_timeout(100)
    
    # 3. 点击月份输入框
    frame.get_by_role("textbox").first.click()
    page.wait_for_timeout(100)
    
    # 4. 选择月份
    frame.get_by_role("cell", name=month_cn, exact=True).click()
    page.wait_for_timeout(100)
    
    # 5. 选择日期必须精确匹配；否则 9 会匹配到 29，导致页面日期被点到上个月。
    frame.get_by_role("cell", name=day_str, exact=True).first.click()
    page.wait_for_timeout(100)


def query_flight_record(page, emp_id, start_date, end_date, clear_first=False):
    """执行查询操作"""
    # 填写员工号
    emp_input = page.get_by_placeholder("员工号或姓名")
    emp_input.click()
    page.wait_for_timeout(100)
    
    if clear_first:
        # 统信浏览器：先清空再输入
        emp_input.fill("")
        page.wait_for_timeout(100)
    
    # 用type模拟逐字输入
    emp_input.type(str(emp_id), delay=50)
    page.wait_for_timeout(300)
    
    # 填写开始日期 - 多次点击确保触发日期选择器
    page.locator("#flyTimeExperience_beginDate").click()
    page.wait_for_timeout(200)
    if clear_first:
        # 第二次及以后需要多点几次
        page.locator("#flyTimeExperience_beginDate").click()
        page.wait_for_timeout(200)
    fill_date(page, start_date)
    
    # 填写结束日期 - 多次点击确保触发日期选择器
    page.locator("#flyTimeExperience_endDate").click()
    page.wait_for_timeout(200)
    if clear_first:
        # 第二次及以后需要多点几次
        page.locator("#flyTimeExperience_endDate").click()
        page.wait_for_timeout(200)
    fill_date(page, end_date)
    
    # 点击查询
    page.wait_for_timeout(300)
    page.get_by_role("button", name="查询").click()
    page.wait_for_timeout(1500)


def extract_flight_data(page, exp_col_index: int, exp_label: str):
    """提取经历数据"""
    try:
        # 等待表格加载
        page.wait_for_timeout(1000)
        
        # 查找表格tbody
        tbody = page.locator("tbody.list")
        
        # 获取第一行数据
        first_row = tbody.locator("tr").first
        
        # 提取所有td单元格
        cells = first_row.locator("td")
        cell_count = cells.count()
        
        if cell_count == 0:
            return None
        
        # 提取经历列（可选）和起落总数（第16列）
        data = {
            exp_label: cells.nth(exp_col_index).inner_text().strip() if cell_count > exp_col_index else "",
            "起落总数": cells.nth(15).inner_text().strip() if cell_count > 15 else "",
        }
        
        return data
    except Exception as e:
        print(c_warn(f"提取数据失败: {e}"))
        return None


def validate_page_query_dates(page, expected_start: str, expected_end: str) -> tuple:
    """查询后校验页面实际日期，防止日期控件点错后继续批量查询"""
    actual_start = page.locator("#flyTimeExperience_beginDate").input_value()
    actual_end = page.locator("#flyTimeExperience_endDate").input_value()

    start_ok = same_query_date(actual_start, expected_start)
    end_ok = same_query_date(actual_end, expected_end)
    if start_ok and end_ok:
        return True, ""

    return False, (
        f"页面日期校验失败：期望 {expected_start}~{expected_end}，"
        f"实际 {actual_start}~{actual_end}。已停止后续查询。"
    )


def create_output_file(records, exp_label: str) -> str:
    """为粘贴模式创建新的Excel输出文件"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"{exp_label}查询结果_{timestamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.append(["员工号", "姓名", "起始时间", "截止时间", exp_label, "起落总数"])

    for i, record in enumerate(records, start=2):
        ws.cell(row=i, column=1, value=record["员工号"])
        ws.cell(row=i, column=2, value=record.get("姓名", ""))
        write_excel_date_cell(ws, i, 3, record.get("开始日期值"))
        write_excel_date_cell(ws, i, 4, record.get("结束日期值"))
        record["行号"] = i

    wb.save(output_file)
    wb.close()
    return output_file


def write_failed_records_log(failed_records: list, exp_label: str) -> None:
    """把失败记录单独写成文本，便于复制后重跑"""
    if not failed_records:
        return

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{exp_label}查询失败记录_{timestamp}.txt"
    try:
        with open(filename, "w", encoding="utf-8") as file:
            file.write(f"{exp_label}查询失败记录 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            file.write(f"共 {len(failed_records)} 条\n")
            file.write("-" * 80 + "\n")
            for record, reason in failed_records:
                file.write(
                    f"{record.get('员工号', '')}\t"
                    f"{record.get('姓名', '')}\t"
                    f"{record.get('开始日期', '')}\t"
                    f"{record.get('结束日期', '')}\t"
                    f"{reason}\n"
                )
        print(c_warn(f"失败记录已保存: {os.path.abspath(filename)}"))
    except Exception as e:
        print(c_warn(f"保存失败记录失败: {e}"))


def collect_query_batch(exp_label: str, input_mode: str) -> tuple:
    """根据输入方式收集一轮查询数据，返回(status, records, output_file)"""
    records = []
    output_file = None

    if input_mode == "1":
        filepath = input(c_hint("请输入Excel文件路径(输入q退出当前工具): ")).strip().strip('"').strip("'")
        if filepath.lower() == "q":
            return "quit", None, None
        if not filepath:
            print(c_warn("未输入文件路径"))
            return "retry", None, None
        if not os.path.exists(filepath):
            print(c_err("文件不存在"))
            return "retry", None, None

        print(c_info("正在读取Excel文件..."))
        records, errors = parse_excel_file(filepath)
        output_file = filepath

        if errors:
            print(c_err("解析错误:"))
            for err in errors:
                print(c_err(f"  {err}"))

        return "ok", records, output_file

    if input_mode == "2":
        text = read_multiline(c_hint("请粘贴数据(输入ok确认,c取消):"), "ok", "c")
        if text is None:
            print(c_warn("已取消本轮输入"))
            return "retry", None, None

        print(c_info("正在解析数据..."))
        records, errors = parse_batch_input(text)

        if errors:
            print(c_err("解析错误:"))
            for err in errors:
                print(c_err(err))

        if records:
            output_file = create_output_file(records, exp_label)
            print(c_ok(f"已创建输出文件: {output_file}"))

        return "ok", records, output_file

    print(c_err("无效输入方式"))
    return "retry", None, None


def open_flight_record_page(page):
    """进入飞行经历查询页面并回到初始查询状态"""
    try:
        print(c_info("正在进入飞行经历查询页面..."))
        page.goto("https://ieb.csair.com/index/index")
        page.wait_for_load_state("networkidle")
        page.get_by_role("listitem").filter(has_text="统计应用").locator("span").click()
        page.get_by_role("link", name="综合报表").wait_for()
        page.get_by_role("link", name="综合报表").click()
        page.get_by_role("link", name="飞行经历").wait_for()
        page.get_by_role("link", name="飞行经历").click()
        page.wait_for_load_state("networkidle")
        page.wait_for_timeout(500)
        page.get_by_role("radio").nth(2).check()
        page.wait_for_timeout(300)
        print(c_ok("已进入飞行经历查询页面"))
    except Exception as e:
        print(c_err(f"自动导航失败: {e}"))
        print(c_warn("请手动进入飞行经历查询页面"))
        input(c_hint("准备好后按回车继续..."))


def run_batch_query(page, records, output_file: str, exp_label: str, exp_col_index: int) -> tuple:
    """执行一轮批量查询并返回成功/失败数"""
    print(c_ok("开始批量查询"))
    success_count = 0
    fail_count = 0
    failed_records = []

    for i, record in enumerate(records):
        name = record.get("姓名", "未知")
        print(f"{c_info(f'[{i+1}/{len(records)}]')} 查询: {record['员工号']} {name} {record['开始日期']}~{record['结束日期']}")

        try:
            query_flight_record(page, record["员工号"], record["开始日期"], record["结束日期"], clear_first=(i > 0))
            date_ok, date_reason = validate_page_query_dates(page, record["开始日期"], record["结束日期"])
            if not date_ok:
                print(c_err(date_reason))
                write_to_excel(
                    output_file,
                    record["行号"],
                    "日期校验失败",
                    "日期校验失败",
                    exp_label,
                    record.get("开始日期值"),
                    record.get("结束日期值"),
                )
                failed_records.append((record, date_reason))
                fail_count += 1
                break

            data = extract_flight_data(page, exp_col_index, exp_label)

            if data and data.get(exp_label) and data.get("起落总数"):
                exp_value = data[exp_label]
                landing_count = data["起落总数"]

                if write_to_excel(
                    output_file,
                    record["行号"],
                    exp_value,
                    landing_count,
                    exp_label,
                    record.get("开始日期值"),
                    record.get("结束日期值"),
                ):
                    print(c_ok(f"✓ {exp_label}: {exp_value} | 起落总数: {landing_count} | 已写入Excel"))
                    success_count += 1
                else:
                    print(c_err("✗ 查询成功但写入失败"))
                    failed_records.append((record, "查询成功但写入Excel失败"))
                    fail_count += 1
            else:
                print(c_warn("✗ 未查询到数据"))
                write_to_excel(
                    output_file,
                    record["行号"],
                    "无数据",
                    "无数据",
                    exp_label,
                    record.get("开始日期值"),
                    record.get("结束日期值"),
                )
                failed_records.append((record, "未查询到数据"))
                fail_count += 1

        except Exception as e:
            print(c_err(f"✗ 失败: {e}"))
            write_to_excel(
                output_file,
                record["行号"],
                "查询失败",
                "查询失败",
                exp_label,
                record.get("开始日期值"),
                record.get("结束日期值"),
            )
            failed_records.append((record, str(e)))
            fail_count += 1

    write_failed_records_log(failed_records, exp_label)
    return success_count, fail_count


def ask_next_action(prompt: str) -> str:
    """批次结束后的下一步动作：继续/切换输入方式/切换查询类型/退出"""
    action = input(c_hint(prompt)).strip().lower()
    if action in {"", "m", "t", "q"}:
        return action
    print(c_warn("未识别输入，默认继续当前模式"))
    return ""


def main():
    print(c_info("飞行经历/左座经历起落数查询助手 - 批量模式（统信浏览器版本）"))
    
    # 启动浏览器
    print(c_info("正在启动浏览器..."))
    pw = sync_playwright().start()
    browser = pw.chromium.launch(headless=False)
    context = browser.new_context()
    context.set_default_timeout(0)
    page = context.new_page()
    
    try:
        try:
            # 登录
            page.goto("https://ieb.csair.com/login")
            page.wait_for_load_state("networkidle")
            page.locator("#scanLogin").wait_for()
            page.locator("#scanLogin").click()
            print(c_info("请扫码登录..."))
            page.wait_for_url("**/index/**")
            page.wait_for_load_state("networkidle")
            print(c_ok("登录成功"))
        except Exception as e:
            print(c_err(f"自动登录失败: {e}"))
            print(c_warn("请手动完成登录"))
            input(c_hint("登录完成后按回车继续..."))

        open_flight_record_page(page)
        print(c_ok("开始工作"))
        
        current_query = select_query_type()
        if current_query is None:
            return

        current_input_mode = select_input_mode()
        if current_input_mode is None:
            return

        while True:
            exp_label, exp_col_index = current_query
            print(c_ok(f"\n当前查询类型: {exp_label}"))
            print(c_ok(f"当前输入方式: {INPUT_MODE_OPTIONS[current_input_mode]}"))

            status, records, output_file = collect_query_batch(exp_label, current_input_mode)

            if status == "quit":
                return

            if status != "ok" or not records:
                if status == "ok":
                    print(c_err("没有可处理的记录"))

                action = ask_next_action("按回车回到当前输入区，输入m切换输入方式，输入t切换查询类型，输入q退出: ")
                if action == "q":
                    return
                if action == "t":
                    current_query = select_query_type(current_query)
                    if current_query is None:
                        return
                if action == "m":
                    current_input_mode = select_input_mode(current_input_mode)
                    if current_input_mode is None:
                        return
                continue

            print(c_ok(f"共{len(records)}条有效数据"))
            for i, r in enumerate(records, 1):
                name = r.get("姓名", "未知")
                print(f"{i}. {r['员工号']} {name} {r['开始日期']}~{r['结束日期']}")

            confirm = ask_next_action("按回车开始查询，输入m切换输入方式，输入t切换查询类型，输入q退出: ")
            if confirm == "q":
                return
            if confirm == "t":
                current_query = select_query_type(current_query)
                if current_query is None:
                    return
                continue
            if confirm == "m":
                current_input_mode = select_input_mode(current_input_mode)
                if current_input_mode is None:
                    return
                continue

            success_count, fail_count = run_batch_query(page, records, output_file, exp_label, exp_col_index)

            print(c_ok(f"\n批量查询完成！成功: {success_count}, 失败: {fail_count}"))
            print(c_ok(f"结果已保存到: {os.path.abspath(output_file)}"))
            print(c_info("\n浏览器保持打开状态，可以手动查看结果"))

            action = ask_next_action("按回车继续当前输入方式查询下一批，输入m切换输入方式，输入t切换查询类型，输入q退出程序: ")
            if action == "q":
                return
            if action == "t":
                current_query = select_query_type(current_query)
                if current_query is None:
                    return
            if action == "m":
                current_input_mode = select_input_mode(current_input_mode)
                if current_input_mode is None:
                    return
    finally:
        try:
            context.close()
        except Exception:
            pass
        try:
            browser.close()
        except Exception:
            pass
        try:
            pw.stop()
        except Exception:
            pass


if __name__ == "__main__":
    main()
