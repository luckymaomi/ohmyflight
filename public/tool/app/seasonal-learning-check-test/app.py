from __future__ import annotations

import argparse
import re
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from playwright.sync_api import Error, Page, TimeoutError, sync_playwright


LOGIN_URL = "https://ieb.csair.com/login"
INDEX_URL = "https://ieb.csair.com/index/index"
SYSTEM_TRAINING_NAME = "2026年飞行系统夏秋换季学习"
GENERIC_TRAINING_NAME = "换季学习"
EXPECTED_HOURS = "4"
EXPECTED_LOCATION = "飞行学堂"
EXPECTED_OPERATOR = "SYSTEM"
CATEGORY_A_NAME = "类别A：2026年飞行系统夏秋换季学习"
CATEGORY_B_NAME = "类别B：换季学习"
STATUS_NORMAL = "正常"
STATUS_HOURS_MISMATCH = "课时不符"
STATUS_NOT_FOUND = "未查到记录"
STATUS_TIMEOUT = "查询超时"
STATUS_PAGE_ERROR = "页面异常"
EXCEL_HEADERS = ["员工号", "姓名", "匹配类别", "培训项目", "培训课时", "培训地点", "经办人", "状态", "备注", "核对时间"]


@dataclass
class EmployeeRecord:
    employee_id: str
    name: str = ""
    row_num: int = 0


@dataclass
class CheckResult:
    employee_id: str
    name: str
    match_category: str
    training_name: str
    hours: str | None
    location: str
    operator: str
    status: str
    note: str
    checked_at: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="培训课时核对脚本：浏览器登录后批量粘贴员工号姓名，逐条写入 Excel。"
    )
    parser.add_argument(
        "inputs",
        nargs="*",
        help="可选：第一批待查询的文本内容。通常留空，登录后再粘贴。",
    )
    parser.add_argument(
        "--login-timeout",
        type=int,
        default=300,
        help="扫码登录最长等待秒数，默认 300。",
    )
    parser.add_argument(
        "--output",
        default="",
        help="输出 Excel 文件名。若连续多批，后续批次会自动追加 batch 后缀。",
    )
    parser.add_argument(
        "--reopen-after-failures",
        type=int,
        default=3,
        help="连续失败多少次后重新进入主页模块，默认 3。",
    )
    parser.add_argument(
        "--slow-mo",
        type=int,
        default=100,
        help="Playwright slow_mo 毫秒数，默认 100。",
    )
    return parser.parse_args()


def make_output_path(raw_path: str, batch_index: int) -> Path:
    if raw_path:
        base = Path(raw_path)
        if base.suffix.lower() != ".xlsx":
            base = base.with_suffix(".xlsx")
        if batch_index == 1:
            return base.resolve()
        return base.with_name(f"{base.stem}_batch{batch_index:02d}{base.suffix}").resolve()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return Path(f"training_hours_results_{timestamp}_batch{batch_index:02d}.xlsx").resolve()


def read_multiline_batch() -> tuple[str, str | None]:
    print("\n请粘贴员工号和姓名列表，每行一条。输入 ok 开始这一批，输入 c 取消本次输入，输入 q 退出程序：")
    lines: list[str] = []

    while True:
        line = input()
        lowered = line.strip().lower()

        if lowered == "q" and not lines:
            return "quit", None
        if lowered == "c":
            return "cancel", None
        if lowered == "ok":
            break
        if line.strip():
            lines.append(line.rstrip())

    if not lines:
        return "empty", None
    return "ok", "\n".join(lines)


def ask_next_action(prompt: str) -> str:
    action = input(prompt).strip().lower()
    if action in {"", "q"}:
        return action
    print("未识别输入，默认继续。")
    return ""


def parse_single_record(text: str) -> EmployeeRecord | None:
    match = re.search(r"\b(\d{6})\b", text)
    if not match:
        return None

    employee_id = match.group(1)
    remainder = text[match.end() :].strip()
    name = ""
    if remainder:
        name_match = re.search(r"[\u4e00-\u9fa5A-Za-z]+", remainder)
        if name_match:
            name = name_match.group(0)

    return EmployeeRecord(employee_id=employee_id, name=name)


def split_continuous_text(text: str) -> list[str]:
    parts = re.split(r"(?=\d{6}\s*[\u4e00-\u9fa5A-Za-z])", text)
    return [part.strip() for part in parts if part.strip() and re.search(r"\d{6}", part)]


def parse_batch_input(text: str) -> tuple[list[EmployeeRecord], list[str]]:
    records: list[EmployeeRecord] = []
    errors: list[str] = []
    seen: set[str] = set()

    lines = [line.strip() for line in text.strip().splitlines() if line.strip()]
    if len(lines) == 1 and len(lines[0]) > 100:
        lines = split_continuous_text(lines[0])

    for index, line in enumerate(lines, start=1):
        record = parse_single_record(line)
        if not record:
            errors.append(f"第{index}条未识别员工号: {line[:50]}")
            continue
        if record.employee_id in seen:
            errors.append(f"第{index}条员工号重复，已跳过: {record.employee_id}")
            continue
        seen.add(record.employee_id)
        records.append(record)

    return records, errors


def print_records_preview(records: list[EmployeeRecord], errors: list[str]) -> None:
    if errors:
        print("解析提示：")
        for error in errors:
            print(f"  {error}")

    if not records:
        print("没有识别到有效员工记录。")
        return

    print(f"共识别到 {len(records)} 条员工记录：")
    for index, record in enumerate(records, start=1):
        display_name = record.name or "(无姓名)"
        print(f"  {index}. {record.employee_id} {display_name}")


def create_output_workbook(path: Path, records: list[EmployeeRecord]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "核对结果"
    ws.append(EXCEL_HEADERS)

    for index, record in enumerate(records, start=2):
        record.row_num = index
        ws.cell(row=index, column=1, value=record.employee_id)
        ws.cell(row=index, column=2, value=record.name)

    wb.save(path)
    wb.close()


def write_result_to_excel(path: Path, record: EmployeeRecord, result: CheckResult) -> None:
    wb = load_workbook(path)
    ws = wb.active

    ws.cell(row=record.row_num, column=1, value=result.employee_id)
    ws.cell(row=record.row_num, column=2, value=result.name)
    ws.cell(row=record.row_num, column=3, value=result.match_category)
    ws.cell(row=record.row_num, column=4, value=result.training_name)
    ws.cell(row=record.row_num, column=5, value=result.hours or "")
    ws.cell(row=record.row_num, column=6, value=result.location)
    ws.cell(row=record.row_num, column=7, value=result.operator)
    ws.cell(row=record.row_num, column=8, value=result.status)
    ws.cell(row=record.row_num, column=9, value=result.note)
    ws.cell(row=record.row_num, column=10, value=result.checked_at)

    wb.save(path)
    wb.close()


def wait_for_login(page: Page, timeout_seconds: int) -> None:
    print(f"打开登录页: {LOGIN_URL}")
    page.goto(LOGIN_URL)
    page.wait_for_load_state("domcontentloaded")
    page.locator("#scanLogin").wait_for()
    page.locator("#scanLogin").click()
    print(f"请扫码登录，最长等待 {timeout_seconds} 秒...")
    page.wait_for_url("**/index/**", timeout=timeout_seconds * 1000)
    page.wait_for_load_state("networkidle")
    print("登录成功。")


def open_training_record_search(page: Page) -> None:
    page.goto(INDEX_URL)
    page.wait_for_load_state("networkidle")

    qualification_management = page.get_by_text("资质管理").nth(1)
    try:
        qualification_management.wait_for(timeout=3000)
        qualification_management.click()
        page.wait_for_timeout(500)
    except TimeoutError:
        pass

    flight_training = page.get_by_text("飞行训练", exact=True).nth(1)
    flight_training.wait_for()
    flight_training.click()

    technical_docs = page.get_by_role("link", name="技术资料")
    technical_docs.wait_for()
    technical_docs.click()

    data_management = page.get_by_role("link", name="资料管理")
    data_management.wait_for()
    data_management.click()

    search_input = page.get_by_placeholder("员工号或姓名简拼")
    search_input.wait_for()
    print("已进入资料管理查询页。")


def search_employee(page: Page, employee_id: str) -> None:
    search_input = page.get_by_placeholder("员工号或姓名简拼")
    search_input.wait_for(timeout=10000)
    search_input.click()
    page.wait_for_timeout(100)
    search_input.fill("")
    page.wait_for_timeout(100)
    search_input.type(employee_id, delay=50)
    page.wait_for_timeout(300)

    page.get_by_role("button", name="查询").click()
    page.wait_for_timeout(1500)

    employee_link = page.get_by_role("link", name=employee_id)
    employee_link.wait_for(timeout=15000)
    employee_link.click()
    page.wait_for_load_state("networkidle")

    training_record_link = page.get_by_role("link", name="训练检查记录")
    training_record_link.wait_for(timeout=15000)
    training_record_link.click()
    page.wait_for_load_state("networkidle")


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def latest_date_key(fields: dict[str, object]) -> tuple[str, str, int]:
    end_date = str(fields.get("end_date", "") or "")
    start_date = str(fields.get("start_date", "") or "")
    row_index = int(fields.get("row_index", 0) or 0)
    return (end_date or start_date, start_date, row_index)


def expand_target_training(page: Page) -> bool:
    candidates = [
        page.get_by_text(SYSTEM_TRAINING_NAME).first,
        page.get_by_text(GENERIC_TRAINING_NAME).first,
    ]

    for training_entry in candidates:
        try:
            training_entry.wait_for(timeout=3000)
            training_entry.click()
            page.wait_for_timeout(1000)
            return True
        except TimeoutError:
            continue
    return False


def extract_row_fields(row) -> dict[str, object]:
    cells = row.locator("td")
    cell_count = cells.count()
    values = [normalize_text(cells.nth(index).inner_text()) for index in range(cell_count)]

    return {
        "cell_count": cell_count,
        "training_name": values[1] if cell_count > 1 else "",
        "hours": values[3] if cell_count > 3 else "",
        "location": values[4] if cell_count > 4 else "",
        "operator": values[5] if cell_count > 5 else "",
        "teacher": values[6] if cell_count > 6 else "",
        "start_date": values[7] if cell_count > 7 else "",
        "end_date": values[8] if cell_count > 8 else "",
        "result": values[9] if cell_count > 9 else "",
    }


def is_category_a(fields: dict[str, object]) -> bool:
    return (
        str(fields.get("training_name", "")) == SYSTEM_TRAINING_NAME
        and str(fields.get("location", "")) == EXPECTED_LOCATION
        and str(fields.get("operator", "")) == EXPECTED_OPERATOR
    )


def is_category_b(fields: dict[str, object]) -> bool:
    return str(fields.get("training_name", "")) == GENERIC_TRAINING_NAME


def build_result_from_match(
    record: EmployeeRecord,
    match_category: str,
    fields: dict[str, object],
) -> CheckResult:
    hours_text = str(fields.get("hours", "") or "")
    location = str(fields.get("location", "") or "")
    operator = str(fields.get("operator", "") or "")
    training_name = str(fields.get("training_name", "") or "")

    if hours_text == EXPECTED_HOURS:
        status = STATUS_NORMAL
        note = (
            f"命中{match_category}，培训课时为{EXPECTED_HOURS}。"
            f"培训地点：{location or '空'}；经办人：{operator or '空'}。"
        )
    else:
        status = STATUS_HOURS_MISMATCH
        note = (
            f"命中{match_category}，但培训课时为{hours_text or '空'}，应为{EXPECTED_HOURS}。"
            f"培训地点：{location or '空'}；经办人：{operator or '空'}。"
        )

    return CheckResult(
        employee_id=record.employee_id,
        name=record.name,
        match_category=match_category,
        training_name=training_name,
        hours=hours_text or None,
        location=location,
        operator=operator,
        status=status,
        note=note,
        checked_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    )


def extract_training_hours(page: Page, record: EmployeeRecord) -> CheckResult:
    if not expand_target_training(page):
        return CheckResult(
            employee_id=record.employee_id,
            name=record.name,
            match_category="",
            training_name="",
            hours=None,
            location="",
            operator="",
            status=STATUS_NOT_FOUND,
            note="未查找到相关记录",
            checked_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        )

    rows = page.locator("tr[id^='dataRow_']")
    row_count = rows.count()
    if row_count == 0:
        return CheckResult(
            employee_id=record.employee_id,
            name=record.name,
            match_category="",
            training_name="",
            hours=None,
            location="",
            operator="",
            status=STATUS_NOT_FOUND,
            note="未查找到相关记录",
            checked_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        )

    category_a_matches: list[dict[str, object]] = []
    category_b_matches: list[dict[str, object]] = []

    for index in range(row_count):
        row = rows.nth(index)
        fields = extract_row_fields(row)
        fields["row_index"] = index

        if is_category_a(fields):
            category_a_matches.append(fields)
        elif is_category_b(fields):
            category_b_matches.append(fields)

    if category_a_matches:
        selected = max(category_a_matches, key=latest_date_key)
        return build_result_from_match(record, CATEGORY_A_NAME, selected)

    if category_b_matches:
        selected = max(category_b_matches, key=latest_date_key)
        return build_result_from_match(record, CATEGORY_B_NAME, selected)

    return CheckResult(
        employee_id=record.employee_id,
        name=record.name,
        match_category="",
        training_name="",
        hours=None,
        location="",
        operator="",
        status=STATUS_NOT_FOUND,
        note="未查找到相关记录",
        checked_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    )


def close_employee_detail(page: Page) -> bool:
    close_targets = [
        page.locator("#showBasicInfoPage .el-dialog__headerbtn").first,
        page.locator("#showBasicInfoPage .el-dialog__close").first,
        page.locator("div").filter(has_text=re.compile(r"^个人信息close$")).locator("span").nth(2),
    ]

    for target in close_targets:
        try:
            target.click(timeout=1500)
            page.wait_for_timeout(800)
            return True
        except TimeoutError:
            continue
        except Error:
            continue
    return False


def clear_search_input(page: Page) -> bool:
    try:
        search_input = page.get_by_placeholder("员工号或姓名简拼")
        search_input.wait_for(timeout=10000)
        search_input.click()
        page.wait_for_timeout(100)
        search_input.fill("")
        page.wait_for_timeout(200)
        return True
    except TimeoutError:
        return False
    except Error:
        return False


def reset_for_next_employee(page: Page) -> bool:
    close_employee_detail(page)
    return clear_search_input(page)


def build_error_result(record: EmployeeRecord, status: str, note: str) -> CheckResult:
    return CheckResult(
        employee_id=record.employee_id,
        name=record.name,
        match_category="",
        training_name="",
        hours=None,
        location="",
        operator="",
        status=status,
        note=note,
        checked_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    )


def attempt_employee_check(page: Page, record: EmployeeRecord, reopen_after_failures: int) -> CheckResult:
    failure_notes: list[str] = []

    for attempt in range(1, reopen_after_failures + 1):
        try:
            print(f"  当前页尝试 {attempt}/{reopen_after_failures}")
            search_employee(page, record.employee_id)
            result = extract_training_hours(page, record)

            if not reset_for_next_employee(page):
                print("  当前页复位失败，重新进入资料管理页。")
                open_training_record_search(page)

            return result
        except (TimeoutError, Error) as exc:
            failure_note = str(exc).splitlines()[0]
            failure_notes.append(failure_note)
            print(f"  第 {attempt} 次失败: {failure_note}")

            if not reset_for_next_employee(page):
                open_training_record_search(page)

    print(f"  连续失败 {reopen_after_failures} 次，重新进入主页模块后再试一次...")
    open_training_record_search(page)

    try:
        search_employee(page, record.employee_id)
        result = extract_training_hours(page, record)

        if not reset_for_next_employee(page):
            open_training_record_search(page)

        return result
    except (TimeoutError, Error) as exc:
        failure_note = str(exc).splitlines()[0]
        if not reset_for_next_employee(page):
            try:
                open_training_record_search(page)
            except (TimeoutError, Error):
                pass

        if isinstance(exc, TimeoutError):
            status = STATUS_TIMEOUT
            note = "查询超时，自动重试后仍未成功"
        else:
            status = STATUS_PAGE_ERROR
            note = "页面异常，自动重试后仍未成功"
        return build_error_result(record, status, note)


def print_result(result: CheckResult, output_path: Path) -> None:
    hours_text = result.hours if result.hours is not None else "(空)"
    display_name = result.name or "(无姓名)"
    category_text = result.match_category or "未匹配到类别"
    training_text = result.training_name or "未查找到相关记录"
    operator_text = result.operator or "(空)"
    print(
        f"[{result.employee_id} {display_name}] "
        f"{category_text} / {training_text} -> 课时={hours_text}, 经办人={operator_text}, 状态={result.status}, 备注={result.note}"
    )
    print(f"  已写入 Excel: {output_path}")


def run_batch(
    page: Page,
    records: list[EmployeeRecord],
    output_path: Path,
    reopen_after_failures: int,
) -> tuple[int, int, int, int]:
    normal_count = 0
    mismatch_count = 0
    not_found_count = 0
    other_count = 0

    for index, record in enumerate(records, start=1):
        display_name = record.name or "(无姓名)"
        print(f"\n[{index}/{len(records)}] 正在查询 {record.employee_id} {display_name} ...")

        result = attempt_employee_check(
            page,
            record,
            reopen_after_failures=reopen_after_failures,
        )
        write_result_to_excel(output_path, record, result)
        print_result(result, output_path)

        if result.status == STATUS_NORMAL:
            normal_count += 1
        elif result.status == STATUS_HOURS_MISMATCH:
            mismatch_count += 1
        elif result.status == STATUS_NOT_FOUND:
            not_found_count += 1
        else:
            other_count += 1

        time.sleep(0.5)

    return normal_count, mismatch_count, not_found_count, other_count


def collect_first_batch_text(args: argparse.Namespace) -> str | None:
    if not args.inputs:
        return None
    return "\n".join(args.inputs)


def main() -> int:
    args = parse_args()
    pending_text = collect_first_batch_text(args)
    batch_index = 1

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=False, slow_mo=args.slow_mo)
        context = browser.new_context(ignore_https_errors=True)
        context.set_default_timeout(30000)
        page = context.new_page()

        try:
            wait_for_login(page, timeout_seconds=args.login_timeout)
            open_training_record_search(page)

            while True:
                if pending_text is not None:
                    batch_text = pending_text
                    pending_text = None
                else:
                    status, batch_text = read_multiline_batch()
                    if status == "quit":
                        break
                    if status in {"cancel", "empty"}:
                        continue

                records, errors = parse_batch_input(batch_text)
                print_records_preview(records, errors)
                if not records:
                    continue

                confirm = ask_next_action("按回车开始查询这一批，输入 q 退出程序：")
                if confirm == "q":
                    break

                output_path = make_output_path(args.output, batch_index)
                batch_index += 1
                create_output_workbook(output_path, records)
                print(f"已创建输出文件: {output_path}")

                normal_count, mismatch_count, not_found_count, other_count = run_batch(
                    page,
                    records,
                    output_path,
                    reopen_after_failures=args.reopen_after_failures,
                )

                print("\n本批次核对完成。")
                print(f"正常: {normal_count}")
                print(f"课时异常: {mismatch_count}")
                print(f"未查到记录: {not_found_count}")
                print(f"其他异常: {other_count}")
                print(f"结果文件: {output_path}")

                open_training_record_search(page)
                print("\n浏览器保持打开，已回到资料管理查询页。")
                action = ask_next_action("按回车继续下一批，输入 q 退出程序：")
                if action == "q":
                    break
        finally:
            try:
                context.close()
            finally:
                browser.close()

    return 0


if __name__ == "__main__":
    sys.exit(main())
