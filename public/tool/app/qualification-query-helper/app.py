#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
IEB 技术等级/运行资格查询助手。

只支持 Excel 输入：员工号为必填列，姓名为可选列。
每查询完一个员工，立即把明细和处理报告写入结果工作簿。
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from playwright.sync_api import sync_playwright


INPUT_HEADERS = ("员工号", "姓名")
DETAIL_HEADERS = (
    "员工号",
    "姓名",
    "资料类型",
    "页面序号",
    "类型",
    "代码",
    "名称",
    "水平等级",
    "机型",
    "生效时间",
    "失效时间",
    "对应检查记录",
    "数据来源",
    "备注",
    "抓取状态",
    "说明",
    "抓取时间",
)
REPORT_HEADERS = (
    "输入行号",
    "员工号",
    "输入姓名",
    "页面姓名",
    "员工号匹配",
    "姓名匹配",
    "技术等级条数",
    "运行资格条数",
    "抓取状态",
    "说明",
    "查询时间",
)
SUMMARY_HEADERS = ("项目", "值")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
DEFAULT_TEMPLATE_NAME = "技术等级运行资格查询模板.xlsx"


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def normalize_emp_id(value: Any) -> str:
    text = normalize_text(value)
    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0]
    return text if text.isdigit() else ""


def normalize_name(value: Any) -> str:
    text = normalize_text(value)
    text = re.sub(r"[（(][^）)]*[）)]", "", text)
    return re.sub(r"\s+", "", text)


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def default_output_path(input_path: str) -> str:
    path = Path(input_path)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return str(path.with_name(f"{path.stem}_技术等级运行资格查询结果_{timestamp}.xlsx"))


def create_template(path: str) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "查询名单"
    worksheet.append(list(INPUT_HEADERS))
    worksheet.column_dimensions["A"].width = 14
    worksheet.column_dimensions["B"].width = 14
    style_header(worksheet)
    workbook.save(path)
    workbook.close()


def style_header(worksheet) -> None:
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def find_header_indexes(headers: list[str]) -> tuple[int | None, int | None]:
    normalized = [normalize_text(value) for value in headers]
    emp_index = next((index for index, value in enumerate(normalized) if value in {"员工号", "工号"}), None)
    name_index = next((index for index, value in enumerate(normalized) if value == "姓名"), None)
    return emp_index, name_index


def read_input_records(path: str) -> tuple[list[dict], list[dict]]:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    workbook.close()

    if not rows:
        return [], [{"输入行号": 0, "员工号": "", "姓名": "", "说明": "Excel 为空"}]

    emp_index, name_index = find_header_indexes(list(rows[0]))
    if emp_index is None:
        return [], [{"输入行号": 1, "员工号": "", "姓名": "", "说明": "未找到“员工号”或“工号”表头"}]

    records = []
    errors = []
    seen_emp_ids = set()
    for row_number, row in enumerate(rows[1:], start=2):
        if not row or not any(normalize_text(value) for value in row):
            continue
        raw_emp_id = row[emp_index] if emp_index < len(row) else ""
        emp_id = normalize_emp_id(raw_emp_id)
        name = normalize_text(row[name_index]) if name_index is not None and name_index < len(row) else ""

        if not re.fullmatch(r"\d{6}", emp_id):
            errors.append({"输入行号": row_number, "员工号": normalize_text(raw_emp_id), "姓名": name, "说明": "员工号不是 6 位数字"})
            continue
        if emp_id in seen_emp_ids:
            errors.append({"输入行号": row_number, "员工号": emp_id, "姓名": name, "说明": "重复员工号，已跳过"})
            continue

        seen_emp_ids.add(emp_id)
        records.append({"输入行号": row_number, "员工号": emp_id, "姓名": name})

    return records, errors


def create_result_workbook(output_path: str) -> None:
    workbook = Workbook()
    report_sheet = workbook.active
    report_sheet.title = "处理报告"
    detail_sheet = workbook.create_sheet("技术资料明细")
    summary_sheet = workbook.create_sheet("汇总")

    report_sheet.append(list(REPORT_HEADERS))
    detail_sheet.append(list(DETAIL_HEADERS))
    summary_sheet.append(list(SUMMARY_HEADERS))
    for worksheet in (report_sheet, detail_sheet, summary_sheet):
        style_header(worksheet)

    report_sheet.column_dimensions["A"].width = 11
    report_sheet.column_dimensions["B"].width = 12
    report_sheet.column_dimensions["C"].width = 12
    report_sheet.column_dimensions["D"].width = 12
    report_sheet.column_dimensions["E"].width = 12
    report_sheet.column_dimensions["F"].width = 12
    report_sheet.column_dimensions["G"].width = 16
    report_sheet.column_dimensions["H"].width = 16
    report_sheet.column_dimensions["I"].width = 14
    report_sheet.column_dimensions["J"].width = 48
    report_sheet.column_dimensions["K"].width = 20

    for column, width in {
        "A": 12, "B": 12, "C": 12, "D": 10, "E": 18, "F": 16, "G": 34,
        "H": 12, "I": 10, "J": 14, "K": 14, "L": 18, "M": 12, "N": 12,
        "O": 12, "P": 48, "Q": 20,
    }.items():
        detail_sheet.column_dimensions[column].width = width

    summary_sheet.column_dimensions["A"].width = 20
    summary_sheet.column_dimensions["B"].width = 90
    workbook.save(output_path)
    workbook.close()


def append_employee_result(output_path: str, record: dict, result: dict) -> None:
    workbook = load_workbook(output_path)
    report_sheet = workbook["处理报告"]
    detail_sheet = workbook["技术资料明细"]

    page_name = result.get("页面姓名", "")
    input_name = record.get("姓名", "")
    name_match = "未提供" if not input_name else ("是" if normalize_name(input_name) == normalize_name(page_name) else "否")
    detail_rows = result.get("明细", [])
    for row in detail_rows:
        detail_sheet.append(row)

    report_sheet.append([
        record.get("输入行号", ""),
        record.get("员工号", ""),
        input_name,
        page_name,
        "是" if result.get("员工号匹配") else "否",
        name_match,
        result.get("技术等级条数", 0),
        result.get("运行资格条数", 0),
        result.get("状态", "异常"),
        result.get("说明", ""),
        now_text(),
    ])
    workbook.save(output_path)
    workbook.close()


def write_summary(output_path: str, input_path: str, total: int, success: int, failed: int, input_errors: int, interrupted: bool) -> str:
    report_path = str(Path(output_path).with_suffix(".txt"))
    summary_rows = [
        ("输入文件", os.path.abspath(input_path)),
        ("结果文件", os.path.abspath(output_path)),
        ("查询时间", now_text()),
        ("有效员工数", total),
        ("成功人数", success),
        ("失败人数", failed),
        ("输入错误数", input_errors),
        ("是否中断", "是" if interrupted else "否"),
    ]

    workbook = load_workbook(output_path)
    summary_sheet = workbook["汇总"]
    if summary_sheet.max_row > 1:
        summary_sheet.delete_rows(2, summary_sheet.max_row - 1)
    for row in summary_rows:
        summary_sheet.append(row)
    workbook.save(output_path)
    workbook.close()

    with open(report_path, "w", encoding="utf-8") as report_file:
        report_file.write("技术等级运行资格查询报告\n")
        report_file.write("=" * 48 + "\n")
        for key, value in summary_rows:
            report_file.write(f"{key}: {value}\n")
    return report_path


def click_first(page, candidates: list, label: str, timeout: int = 10000) -> None:
    last_error = None
    for locator in candidates:
        try:
            locator.first.wait_for(state="visible", timeout=timeout)
            locator.first.click(timeout=timeout)
            return
        except Exception as exc:
            last_error = exc
    raise RuntimeError(f"未能点击{label}: {last_error}")


def login(page) -> None:
    page.goto("https://ieb.csair.com/login")
    try:
        page.wait_for_load_state("networkidle", timeout=10000)
    except Exception:
        pass
    try:
        page.locator("#scanLogin").wait_for(state="visible", timeout=10000)
        page.locator("#scanLogin").click()
        print("请扫码登录...")
        page.wait_for_url("**/index/**", timeout=25000)
        print("登录成功")
    except Exception as exc:
        print(f"自动等待登录未完成: {exc}")
        input("请手动完成登录，并停留在 IEB 首页后按回车继续: ")


def open_material_management(page) -> None:
    page.goto("https://ieb.csair.com/index/index")
    try:
        page.wait_for_load_state("networkidle", timeout=10000)
    except Exception:
        pass
    click_first(page, [page.get_by_text("资质管理").nth(1), page.get_by_text("资质管理")], "资质管理")
    click_first(page, [page.get_by_text("飞行训练").nth(1), page.get_by_text("飞行训练")], "飞行训练")
    click_first(page, [page.get_by_role("link", name="技术资料"), page.get_by_text("技术资料")], "技术资料")
    click_first(page, [page.get_by_role("link", name="资料管理"), page.get_by_text("资料管理")], "资料管理")
    page.get_by_role("textbox", name="员工号或姓名简拼").wait_for(state="visible", timeout=12000)


def wait_for_search_result(page, emp_id: str, timeout_ms: int = 10000) -> None:
    deadline = datetime.now().timestamp() + timeout_ms / 1000
    while datetime.now().timestamp() < deadline:
        try:
            if page.get_by_role("link", name=emp_id, exact=True).count() > 0:
                return
        except Exception:
            pass
        page.wait_for_timeout(200)
    raise RuntimeError(f"查询后未找到员工号链接: {emp_id}")


def read_name_from_result_table(page, emp_id: str) -> str:
    script = """
    (targetId) => {
      const clean = (value) => (value || '').replace(/\\s+/g, ' ').trim();
      for (const row of Array.from(document.querySelectorAll('tr'))) {
        const values = Array.from(row.querySelectorAll('th,td')).map((cell) => clean(cell.innerText));
        const index = values.findIndex((value) => value === targetId);
        if (index >= 0) {
          const name = values.slice(index + 1).find((value) => /^[\\u4e00-\\u9fa5·]{2,8}$/.test(value));
          if (name) return name;
        }
      }
      return '';
    }
    """
    try:
        return normalize_text(page.evaluate(script, emp_id))
    except Exception:
        return ""


def search_employee(page, emp_id: str) -> str:
    textbox = page.get_by_role("textbox", name="员工号或姓名简拼")
    textbox.click()
    textbox.fill("")
    textbox.type(emp_id, delay=20)
    page.get_by_role("button", name="查询").click()
    wait_for_search_result(page, emp_id)
    page_name = read_name_from_result_table(page, emp_id)
    page.get_by_role("link", name=emp_id, exact=True).click()
    page.wait_for_timeout(600)
    return page_name


EXTRACT_TABLE_JS = """
(container) => {
  const clean = (value) => (value || '').replace(/\\s+/g, ' ').trim();
  const parseTable = (table, expectedCols) => {
    const active = new Map();
    const result = [];
    for (const tr of Array.from(table.querySelectorAll('tr'))) {
      const values = [];
      let column = 0;
      const consumeActive = () => {
        const entry = active.get(column);
        values.push(entry.text);
        if (entry.rowsLeft <= 1) active.delete(column);
        else active.set(column, { text: entry.text, rowsLeft: entry.rowsLeft - 1 });
        column += 1;
      };
      const cells = Array.from(tr.children).filter((cell) => cell.tagName === 'TH' || cell.tagName === 'TD');
      for (const cell of cells) {
        while (active.has(column)) consumeActive();
        const text = clean(cell.innerText);
        const rowspan = Number(cell.getAttribute('rowspan') || 1);
        const colspan = Number(cell.getAttribute('colspan') || 1);
        for (let offset = 0; offset < colspan; offset += 1) {
          values.push(text);
          if (rowspan > 1) active.set(column + offset, { text, rowsLeft: rowspan - 1 });
        }
        column += colspan;
      }
      const limit = expectedCols || Math.max(values.length, ...Array.from(active.keys(), (key) => key + 1), 0);
      while (column < limit) {
        if (active.has(column)) consumeActive();
        else {
          values.push('');
          column += 1;
        }
      }
      if (values.some(Boolean)) result.push(values);
    }
    return result;
  };
  const tables = Array.from(container.querySelectorAll('table'));
  if (tables.length < 2) return { headers: [], rows: [], tableCount: tables.length };
  const headers = parseTable(tables[0], 0)[0] || [];
  const rows = parseTable(tables[1], headers.length);
  return { headers, rows, tableCount: tables.length };
}
"""


def wait_for_container(page, selector: str, timeout_ms: int = 8000) -> None:
    deadline = datetime.now().timestamp() + timeout_ms / 1000
    while datetime.now().timestamp() < deadline:
        for frame in page.frames:
            try:
                if frame.locator(selector).count() > 0:
                    return
            except Exception:
                pass
        page.wait_for_timeout(200)
    raise RuntimeError(f"未找到目标容器: {selector}")


def extract_table(page, selector: str) -> tuple[list[str], list[dict]]:
    wait_for_container(page, selector)
    for frame in page.frames:
        try:
            locator = frame.locator(selector)
            if locator.count() == 0:
                continue
            parsed = locator.first.evaluate(EXTRACT_TABLE_JS)
            headers = [normalize_text(value) for value in parsed.get("headers", [])]
            rows = [
                dict(zip(headers, [normalize_text(value) for value in values[:len(headers)]]))
                for values in parsed.get("rows", [])
                if any(normalize_text(value) for value in values)
            ]
            return headers, rows
        except Exception:
            continue
    raise RuntimeError(f"读取目标容器失败: {selector}")


def read_tab(page, label: str, selector: str, required_headers: set[str]) -> list[dict]:
    candidates = [page.get_by_role("link", name=label, exact=True), page.get_by_text(label, exact=True)]
    last_error = None
    for attempt in range(2):
        try:
            click_first(page, candidates, label)
            headers, rows = extract_table(page, selector)
            if not required_headers.issubset(set(headers)):
                raise RuntimeError(f"{label}表头异常: {headers}")
            return rows
        except Exception as exc:
            last_error = exc
            if attempt == 0:
                page.wait_for_timeout(1000)
    raise RuntimeError(str(last_error))


def close_person_dialog(page) -> None:
    try:
        close_button = page.locator(".pilotInfo-dialog-close")
        if close_button.count() > 0:
            close_button.first.click(timeout=5000)
            page.wait_for_timeout(400)
    except Exception:
        pass


def detail_rows(record: dict, page_name: str, tech_rows: list[dict], operation_rows: list[dict]) -> list[list[str]]:
    name = page_name or record.get("姓名", "")
    captured_at = now_text()
    rows = []
    for row in tech_rows:
        rows.append([
            record["员工号"], name, "技术等级", row.get("#", ""), "", row.get("技术等级代码", ""),
            row.get("技术等级", ""), row.get("水平等级", ""), row.get("机型", ""),
            row.get("生效时间", ""), row.get("失效时间", ""), row.get("对应检查记录", ""),
            row.get("数据来源", ""), "", "成功", "", captured_at,
        ])
    for index, row in enumerate(operation_rows, start=1):
        rows.append([
            record["员工号"], name, "运行资格", str(index), row.get("类型", ""), row.get("运行资格代码", ""),
            row.get("运行资格", ""), row.get("水平等级", ""), row.get("机型", ""),
            row.get("生效时间", ""), row.get("失效时间", ""), "", "",
            row.get("备注", ""), "成功", "", captured_at,
        ])
    return rows


def query_one(page, record: dict) -> dict:
    page_name = search_employee(page, record["员工号"])
    try:
        tech_rows = read_tab(
            page,
            "技术等级",
            "#qualList",
            {"#", "技术等级代码", "技术等级", "水平等级", "机型", "生效时间", "失效时间"},
        )
        operation_rows = read_tab(
            page,
            "运行资格",
            "#showSingleEmpOperQualList",
            {"类型", "运行资格代码", "运行资格", "水平等级", "机型", "生效时间", "失效时间", "备注"},
        )
        return {
            "状态": "成功",
            "说明": "",
            "页面姓名": page_name,
            "员工号匹配": True,
            "技术等级条数": len(tech_rows),
            "运行资格条数": len(operation_rows),
            "明细": detail_rows(record, page_name, tech_rows, operation_rows),
        }
    finally:
        close_person_dialog(page)


def append_input_error(output_path: str, error: dict) -> None:
    append_employee_result(output_path, error, {
        "状态": "输入错误",
        "说明": error.get("说明", ""),
        "页面姓名": "",
        "员工号匹配": False,
        "技术等级条数": 0,
        "运行资格条数": 0,
        "明细": [],
    })


def is_browser_closed_error(error: Exception) -> bool:
    text = str(error)
    return "Target page, context or browser has been closed" in text or "Browser has been closed" in text


def run_query(input_path: str | None = None, output_path: str | None = None, limit: int = 0) -> int:
    success = 0
    failed = 0
    interrupted = False
    records: list[dict] = []
    input_errors: list[dict] = []
    input_error = ""
    page_ready = False

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=False)
        context = browser.new_context()
        context.set_default_timeout(15000)
        page = context.new_page()
        try:
            login(page)
            open_material_management(page)
            page_ready = True
            print("资料管理页面已就绪。")
            if not input_path:
                input_path = input("请输入 Excel 文件路径: ").strip().strip('"').strip("'")
            if not input_path:
                input_error = "未输入 Excel 文件路径。"
            elif not os.path.exists(input_path):
                input_error = f"文件不存在: {input_path}"
            else:
                output_path = output_path or default_output_path(input_path)
                print(f"输入文件: {os.path.abspath(input_path)}")
                print(f"结果文件: {os.path.abspath(output_path)}")
                print("正在读取 Excel 名单...")
                records, input_errors = read_input_records(input_path)
                if limit:
                    records = records[:limit]
                if not records:
                    print("未识别到可查询的员工号。")
                else:
                    create_result_workbook(output_path)
                    for error in input_errors:
                        append_input_error(output_path, error)
                    print(f"待查询人数: {len(records)}")
                    for index, record in enumerate(records, start=1):
                        print(f"[{index}/{len(records)}] 查询 {record['员工号']} {record.get('姓名', '')}")
                        try:
                            result = query_one(page, record)
                            append_employee_result(output_path, record, result)
                            success += 1
                            print(f"OK 技术等级 {result['技术等级条数']} 条，运行资格 {result['运行资格条数']} 条，已保存")
                        except KeyboardInterrupt:
                            interrupted = True
                            print("收到中断，已停止后续查询。")
                            break
                        except Exception as exc:
                            result = {
                                "状态": "失败",
                                "说明": str(exc),
                                "页面姓名": "",
                                "员工号匹配": False,
                                "技术等级条数": 0,
                                "运行资格条数": 0,
                                "明细": [],
                            }
                            append_employee_result(output_path, record, result)
                            failed += 1
                            print(f"ERROR {record['员工号']}: {exc}")
                            close_person_dialog(page)
                            if is_browser_closed_error(exc):
                                interrupted = True
                                print("浏览器已关闭，已停止后续查询，避免写入无效失败记录。")
                                break
        except KeyboardInterrupt:
            interrupted = True
            print("收到中断，已停止后续查询。")
        except Exception as exc:
            print(f"无法进入资料管理或读取 Excel：{exc}")
        finally:
            try:
                context.close()
            except Exception:
                pass
            try:
                browser.close()
            except Exception:
                pass

    if not page_ready:
        return 2 if interrupted else 3
    if input_error:
        print(input_error)
        return 1
    if not records:
        return 1

    assert input_path is not None
    assert output_path is not None
    report_path = write_summary(
        output_path,
        input_path,
        total=len(records),
        success=success,
        failed=failed,
        input_errors=len(input_errors),
        interrupted=interrupted,
    )
    print(f"完成。成功 {success} 人，失败 {failed} 人。")
    print(f"Excel 结果: {os.path.abspath(output_path)}")
    print(f"文本报告: {os.path.abspath(report_path)}")
    return 0 if failed == 0 and not interrupted else 2


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="IEB 技术等级/运行资格查询助手（仅 Excel 输入）")
    parser.add_argument("--input", help="Excel 输入文件，必须包含“员工号”列，可选“姓名”列")
    parser.add_argument("--output", help="结果 Excel 路径，默认在输入文件同目录生成带时间戳文件")
    parser.add_argument("--limit", type=int, default=0, help="仅查询前 N 人，用于小样本验证")
    parser.add_argument("--create-template", nargs="?", const=DEFAULT_TEMPLATE_NAME, help="创建 Excel 查询模板后退出")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.create_template:
        create_template(args.create_template)
        print(f"模板已创建: {os.path.abspath(args.create_template)}")
        return 0

    return run_query(args.input, args.output, args.limit)


if __name__ == "__main__":
    sys.exit(main())
