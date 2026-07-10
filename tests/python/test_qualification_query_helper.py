import importlib.util
import tempfile
import unittest
from pathlib import Path
from unittest.mock import MagicMock, patch

from openpyxl import Workbook, load_workbook


APP_PATH = (
    Path(__file__).resolve().parents[2]
    / "public"
    / "tool"
    / "app"
    / "qualification-query-helper"
    / "app.py"
)
DOWNLOAD_TEMPLATE_PATH = APP_PATH.with_name("template.xlsx")
SPEC = importlib.util.spec_from_file_location("qualification_query_helper", APP_PATH)
APP = importlib.util.module_from_spec(SPEC)
assert SPEC.loader is not None
SPEC.loader.exec_module(APP)


class QualificationQueryHelperTest(unittest.TestCase):
    def test_template_has_required_headers(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "template.xlsx"
            APP.create_template(str(path))

            workbook = load_workbook(path, data_only=True)
            worksheet = workbook.active
            self.assertEqual(worksheet.title, "查询名单")
            self.assertEqual([cell.value for cell in worksheet[1]], ["员工号", "姓名"])
            workbook.close()

    def test_download_template_has_required_headers(self):
        workbook = load_workbook(DOWNLOAD_TEMPLATE_PATH, data_only=True)
        worksheet = workbook.active
        self.assertEqual(worksheet.title, "查询名单")
        self.assertEqual([cell.value for cell in worksheet[1]], ["员工号", "姓名"])
        workbook.close()

    def test_input_parsing_rejects_invalid_and_duplicate_employee_ids(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "input.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["工号", "姓名"])
            worksheet.append([181558, "赵铁锋"])
            worksheet.append(["193391.0", "陈良"])
            worksheet.append(["A209232", "李杰"])
            worksheet.append(["181558", "重复"])
            workbook.save(path)
            workbook.close()

            records, errors = APP.read_input_records(str(path))

            self.assertEqual(
                records,
                [
                    {"输入行号": 2, "员工号": "181558", "姓名": "赵铁锋"},
                    {"输入行号": 3, "员工号": "193391", "姓名": "陈良"},
                ],
            )
            self.assertEqual(
                [error["说明"] for error in errors],
                ["员工号不是 6 位数字", "重复员工号，已跳过"],
            )

    def test_result_workbook_contains_incremental_report_and_summary(self):
        with tempfile.TemporaryDirectory() as directory:
            output_path = Path(directory) / "result.xlsx"
            APP.create_result_workbook(str(output_path))
            record = {"输入行号": 2, "员工号": "181558", "姓名": "赵铁锋"}
            APP.append_employee_result(
                str(output_path),
                record,
                {
                    "状态": "成功",
                    "说明": "",
                    "页面姓名": "赵铁锋",
                    "员工号匹配": True,
                    "技术等级条数": 1,
                    "运行资格条数": 1,
                    "明细": [
                        ["181558", "赵铁锋", "技术等级", "1", "", "CAPZ", "Z类机长", "", "777",
                         "2021-08-12", "", "", "", "", "成功", "", "2026-07-10 09:00:00"],
                        ["181558", "赵铁锋", "运行资格", "1", "区域航线资格", "RSEA", "欧洲区域单飞资格",
                         "", "777", "2023-10-21", "", "", "", "", "成功", "", "2026-07-10 09:00:00"],
                    ],
                },
            )
            report_path = APP.write_summary(
                str(output_path),
                "input.xlsx",
                total=1,
                success=1,
                failed=0,
                input_errors=0,
                interrupted=False,
            )

            workbook = load_workbook(output_path, data_only=True)
            self.assertEqual(workbook.sheetnames, ["处理报告", "技术资料明细", "汇总"])
            self.assertEqual(workbook["处理报告"].max_row, 2)
            self.assertEqual(workbook["技术资料明细"].max_row, 3)
            self.assertEqual(workbook["汇总"]["B5"].value, 1)
            workbook.close()
            self.assertIn("成功人数: 1", Path(report_path).read_text(encoding="utf-8"))

    def test_interrupted_login_does_not_open_excel_or_create_result_file(self):
        with tempfile.TemporaryDirectory() as directory:
            input_path = Path(directory) / "input.xlsx"
            output_path = Path(directory) / "result.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["员工号", "姓名"])
            worksheet.append(["181558", "赵铁锋"])
            workbook.save(input_path)
            workbook.close()

            playwright = MagicMock()
            browser = MagicMock()
            context = MagicMock()
            page = MagicMock()
            playwright.chromium.launch.return_value = browser
            browser.new_context.return_value = context
            context.new_page.return_value = page
            manager = MagicMock()
            manager.__enter__.return_value = playwright

            with (
                patch.object(APP, "sync_playwright", return_value=manager),
                patch.object(APP, "login", side_effect=KeyboardInterrupt),
                patch.object(APP, "read_input_records") as read_input_records,
            ):
                result = APP.run_query(str(input_path), str(output_path))

            self.assertEqual(result, 2)
            read_input_records.assert_not_called()
            self.assertFalse(output_path.exists())

    def test_excel_is_read_only_after_material_management_is_ready(self):
        with tempfile.TemporaryDirectory() as directory:
            input_path = Path(directory) / "input.xlsx"
            output_path = Path(directory) / "result.xlsx"
            input_path.touch()
            events = []
            playwright = MagicMock()
            browser = MagicMock()
            context = MagicMock()
            page = MagicMock()
            playwright.chromium.launch.return_value = browser
            browser.new_context.return_value = context
            context.new_page.return_value = page
            manager = MagicMock()
            manager.__enter__.return_value = playwright

            def mark_login(_page):
                events.append("login")

            def mark_page_ready(_page):
                events.append("page_ready")

            def read_empty_records(_path):
                events.append("read_excel")
                return [], []

            with (
                patch.object(APP, "sync_playwright", return_value=manager),
                patch.object(APP, "login", side_effect=mark_login),
                patch.object(APP, "open_material_management", side_effect=mark_page_ready),
                patch.object(APP, "read_input_records", side_effect=read_empty_records),
            ):
                result = APP.run_query(str(input_path), str(output_path))

            self.assertEqual(result, 1)
            self.assertEqual(events, ["login", "page_ready", "read_excel"])
            self.assertFalse(output_path.exists())

    def test_excel_path_is_prompted_only_after_material_management_is_ready(self):
        with tempfile.TemporaryDirectory() as directory:
            events = []
            playwright = MagicMock()
            browser = MagicMock()
            context = MagicMock()
            page = MagicMock()
            playwright.chromium.launch.return_value = browser
            browser.new_context.return_value = context
            context.new_page.return_value = page
            manager = MagicMock()
            manager.__enter__.return_value = playwright

            def mark_login(_page):
                events.append("login")

            def mark_page_ready(_page):
                events.append("page_ready")

            def prompt_path(_message):
                events.append("prompt_excel")
                return ""

            with (
                patch.object(APP, "sync_playwright", return_value=manager),
                patch.object(APP, "login", side_effect=mark_login),
                patch.object(APP, "open_material_management", side_effect=mark_page_ready),
                patch("builtins.input", side_effect=prompt_path),
            ):
                result = APP.run_query()

            self.assertEqual(result, 1)
            self.assertEqual(events, ["login", "page_ready", "prompt_excel"])


if __name__ == "__main__":
    unittest.main()
